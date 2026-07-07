#!/usr/bin/env python3
"""
归因脚本 — 五级归因树执行模块

输入：
  1. 主数据文件（Sheet=主数据）— 来自 IM 归因模块输出
  2. 促销数据文件（Sheet1=Q促销分行表, Sheet2=竞品促销分行表）— 来自促销校验模块输出
  3. state.json — 读取 case_audit（每 case 的人工校验原因）
输出：
  将五级归因结果写回主数据文件，同时写入 im一级推导、是否需要人工校验、人工校验分类、{5类}_详情 列；
  追加 Q促销分行表 + 竞品促销分行表 两个 sheet；生成 归因结果汇总（4版本递进）+ 归因逻辑 两个 sheet。

字段名通过 shared/field_map.json 配置，详见 field_definitions.md（shared/）。

用法：
  python3 attribute.py --output-file <主数据.xlsx> --promo-file <促销数据.xlsx> --state-file <state.json>
"""

# ============================================================
# ⚠️ 修改联动声明
# 本脚本与以下文件强绑定（均在 orion-bijiaups-rca-attribution/ 目录）：
#   - attribution_logic.md：共用归因逻辑、携程价格归因、同程价格归因
#   - validation_logic.md ：V3 数据一致性校验规则、人工校验触发规则、路由冲突检测
#
# 改归因规则 → 先更新 attribution_logic.md / validation_logic.md → 再修改本脚本
# 改代码     → 先确认上述 md 中规则描述 → 修改后同步更新对应 md
# 两者不同步 = 归因结果不可信
# ============================================================

import argparse
import json
import re
import sys
import os
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd


# ═══════════════════════════════════════════
# 0. 工具函数
# ═══════════════════════════════════════════

VALID_PLATFORMS = {"携程", "美团", "飞猪", "同程"}
MULTI_VALUE_SEP = "，"
PLATFORM_ALIASES = {
    "xiecheng": "携程",
    "ctrip": "携程",
    "携程": "携程",
    "tongcheng": "同程",
    "tc": "同程",
    "同程": "同程",
}
RULE_SET_BY_PLATFORM = {
    "携程": "xiecheng",
    "同程": "tongcheng",
}
RULE_ROOT = Path("/Users/zhangwang/.claude/agents/orion-bijiaups-rca-shared")

# ── 字段名映射（运行时由 shared/field_map.json 覆盖） ──
# 修改字段名请同步更新 shared/field_map.json 和 shared/field_definitions.md
FIELDS: dict = {
    # 共用输入字段
    "kf_note":          "客服判断场景",
    "compare_platform": "比价平台",
    "q_list_price":     "Q划线价——同质化房型",
    "q_pay_price":      "Q到手价——同质化房型",
    "c_list_price":     "划线价（竞品）",
    "c_pay_price":      "到手价（竞品）",
    "q_promo_detail":   "Q优惠明细",
    "c_promo_detail":   "促销明细（竞品）",
    "case_no":          "case编号",
    "order_no":         "工单号",
    # 平台专用（同程）
    "multi_points":     "是否多倍积分酒店",
    # 平台专用（携程）
    "roomid_q":         "同质化房型roomid",
    "roomid_c":         "CRoomID",
    "is_fenxiao":       "是否分销",
    "is_louchu":        "是否露出",
    "lingjuan":         "领用券情况",
}


def load_field_map(promo_code_set: str) -> dict:
    """从 shared/field_map.json 加载字段名映射（common + 平台专用合并）。"""
    map_path = RULE_ROOT / "field_map.json"
    if not map_path.exists():
        return {}
    with open(map_path, encoding="utf-8") as f:
        data = json.load(f)
    common = data.get("common", {})
    platform_specific = data.get(promo_code_set, {})
    return {**common, **platform_specific}


def _text(row: dict, *field_names: str) -> str:
    for name in field_names:
        value = row.get(name)
        if pd.isna(value):
            continue
        if value not in (None, '', 'nan', 'None'):
            return str(value).strip()
    return ''


def _join_values(values: list) -> str:
    return MULTI_VALUE_SEP.join([str(v) for v in values if str(v or '').strip()])


def normalize_platform(value: str | None) -> str:
    key = str(value or '').strip()
    return PLATFORM_ALIASES.get(key, key)


def require_platform(value: str | None) -> str:
    platform = normalize_platform(value)
    if not platform:
        print("错误：缺少归因平台。当前单次运行必须明确指定平台：携程 / 同程", file=sys.stderr)
        sys.exit(1)
    if platform not in RULE_SET_BY_PLATFORM:
        print(f"错误：暂不支持的归因平台 - {platform}。当前支持：携程 / 同程", file=sys.stderr)
        sys.exit(1)
    return platform


def normalize_rule_set(value: str | None, platform: str) -> str:
    key = str(value or '').strip()
    if key in ("xiecheng", "tongcheng"):
        return key
    return RULE_SET_BY_PLATFORM.get(platform, "xiecheng")


def load_full_state(state_file: str | None) -> dict:
    if not state_file or not os.path.exists(state_file):
        return {}
    with open(state_file, "r", encoding="utf-8") as f:
        return json.load(f)


def _safe_float(v):
    try:
        return float(v) if v not in (None, '', 'nan', 'None') else None
    except (ValueError, TypeError):
        return None


def _case_promos(promo_df: pd.DataFrame, case_id: str, flow_no: str = '') -> pd.DataFrame:
    """先按工单号匹配，找不到再按case编号匹配（OR fallback）。"""
    if promo_df is None or promo_df.empty or 'case编号' not in promo_df.columns:
        return pd.DataFrame()
    if flow_no and '工单号' in promo_df.columns:
        mask = promo_df['工单号'].astype(str) == str(flow_no)
        if mask.any():
            return promo_df[mask]
    return promo_df[promo_df['case编号'].astype(str) == str(case_id)]


def sum_amounts_from_sheet(promo_df: pd.DataFrame, case_id: str, flow_no: str = '', **filters) -> float:
    if promo_df is None or promo_df.empty:
        return 0.0
    subset = _case_promos(promo_df, case_id, flow_no)
    if subset.empty:
        return 0.0
    mask = pd.Series([True] * len(subset), index=subset.index)
    for k, v in filters.items():
        if k in subset.columns:
            mask = mask & (subset[k].astype(str) == str(v))
    return float(subset.loc[mask, '促销金额'].apply(lambda x: _safe_float(x) or 0.0).sum())


def has_promo_name_containing_from_sheet(promo_df: pd.DataFrame, case_id: str, keyword: str, flow_no: str = '') -> bool:
    if promo_df is None or promo_df.empty:
        return False
    subset = _case_promos(promo_df, case_id, flow_no)
    if subset.empty:
        return False
    return any(keyword in str(name) for name in subset['促销名称'])


def _sum_promos(promo_df: pd.DataFrame, case_id: str, predicate, flow_no: str = '') -> float:
    rows = _case_promos(promo_df, case_id, flow_no)
    if rows.empty:
        return 0.0
    mask = rows.apply(predicate, axis=1)
    return float(rows.loc[mask, '促销金额'].apply(lambda x: _safe_float(x) or 0.0).sum())


def _names_promos(promo_df: pd.DataFrame, case_id: str, predicate, flow_no: str = '') -> list:
    rows = _case_promos(promo_df, case_id, flow_no)
    if rows.empty:
        return []
    mask = rows.apply(predicate, axis=1)
    return [str(x) for x in rows.loc[mask, '促销名称'].tolist() if str(x or '').strip()]


# ═══════════════════════════════════════════
# 1. V3 数据一致性校验
# ═══════════════════════════════════════════

def validate_v3_consistency(level1: str, row: dict) -> list:
    errors = []
    q_pay = _safe_float(row.get(FIELDS["q_pay_price"]))
    c_pay = _safe_float(row.get(FIELDS["c_pay_price"]))

    if q_pay is None or c_pay is None:
        return errors

    if level1 == '用户误解' and q_pay > c_pay:
        errors.append('V3错误：用户误解类但Q支付价——同质化房型>到手价（竞品）')
    elif level1 == '价格lose' and q_pay <= c_pay:
        errors.append('V3错误：价格类但Q支付价——同质化房型≤到手价（竞品）')

    return errors


# ═══════════════════════════════════════════
# 2. 一级路由
# ═══════════════════════════════════════════

def route_level1(row: dict) -> str:
    """
    根据客服判断场景路由一级。
    前置：截图平台检查在调用方处理（最高优先级）。
    优先级：异常 > 无法归因 > 库存lose > 用户误解 > 价格lose
    """
    kf = _text(row, FIELDS["kf_note"])

    if not kf:
        return '异常'

    if 'roomid' in kf.lower():
        return '无法归因'
    if '不可对比平台' in kf or '非可对比平台' in kf or '非可比价平台' in kf:
        return '无法归因'
    if '民宿' in kf:
        return '无法归因'
    if '团购' in kf:
        return '无法归因'
    if '预售' in kf:
        return '无法归因'
    if '已下单' in kf:
        return '无法归因'

    if '更好更贵' in kf or '更差更贵' in kf:
        return '库存lose'
    if '物理房型' in kf:
        return '库存lose'
    if '酒店缺失' in kf or '无该酒店' in kf or '没有该酒店' in kf:
        return '库存lose'

    if '酒店不一致' in kf or '入离' in kf or '入住日期' in kf or '入住时间' in kf or '房型不一致' in kf:
        return '用户误解'
    if '价格不lose' in kf or '不lose' in kf or '价格不' in kf:
        return '用户误解'

    if '同质化' in kf:
        return '价格lose'
    if 'lose' in kf:
        return '价格lose'

    return '异常'


# ═══════════════════════════════════════════
# 3. 各分类归因逻辑
# ═══════════════════════════════════════════

def attribute_abnormal(row: dict) -> dict:
    kf = _text(row, FIELDS["kf_note"])
    if not kf:
        return {'一级_AI': '异常', '二级_AI': '信息缺失', '三级_AI': '客服分析结果未填写',
                '四级_AI': '', '五级_AI': '', '归因备注': ''}
    return {'一级_AI': '异常', '二级_AI': '未知原因', '三级_AI': '',
            '四级_AI': '', '五级_AI': '', '归因备注': ''}


def attribute_unattributable(row: dict) -> dict:
    kf = _text(row, FIELDS["kf_note"])
    level2 = ''
    note = ''

    if ('不可对比平台' in kf or '非可对比平台' in kf or '非可比价平台' in kf or
            '民宿' in kf or '团购' in kf or '预售' in kf or '已下单' in kf):
        level2 = '不可对比'
    elif 'roomid' in kf.lower():
        level2 = '无法查询到Croomid'
        c_roomid = str(row.get(FIELDS["roomid_c"], '') or '').strip()
        if not c_roomid or c_roomid in ('nan', 'None'):
            note = 'Croomid为空'

    level3 = ''
    if level2 == '不可对比':
        if '不可对比平台' in kf or '非可对比平台' in kf or '非可比价平台' in kf:
            level3 = '非C平台'
        elif '民宿' in kf:
            level3 = '民宿房'
        elif '团购' in kf:
            level3 = '团购房'
        elif '预售' in kf:
            level3 = '预售'
        elif '已下单' in kf:
            level3 = '已下单'

    return {'一级_AI': '无法归因', '二级_AI': level2, '三级_AI': level3,
            '四级_AI': '', '五级_AI': '', '归因备注': note}


def attribute_inventory(row: dict) -> dict:
    kf = _text(row, FIELDS["kf_note"])
    level2 = ''
    level3 = ''

    if '更好更贵' in kf or '更差更贵' in kf:
        level2 = '同质化房型缺失'
        level3 = '更好更贵' if '更好更贵' in kf else '更差更贵'
    elif '物理房型' in kf:
        level2 = '物理房型缺失'
        if any(x in kf for x in ['已订完', '售罄', '满房', '无房', '订完']):
            level3 = '物理房型已售罄'
        elif any(x in kf for x in ['缺失', '暂无', '无该房型', '没有该房型']):
            level3 = '物理房型真实缺失'
    elif '酒店缺失' in kf or '无该酒店' in kf or '没有该酒店' in kf:
        level2 = '酒店缺失'

    return {'一级_AI': '库存lose', '二级_AI': level2, '三级_AI': level3,
            '四级_AI': '', '五级_AI': '', '归因备注': ''}


def attribute_misconception(row: dict) -> dict:
    im_result = str(row.get('im-归因结果-AI', '') or '').strip()

    hotel_inconsistent = '酒店不一致' in im_result
    inout_inconsistent = '入离不一致' in im_result
    roomtype_inconsistent = '房型不一致' in im_result
    tongzhi_diff = '同质化权益' in im_result
    is_free = '免费权益' in im_result
    is_jifen = '积分' in im_result
    is_cuoluan = '错选乱选' in im_result
    is_unknown = '未知原因' in im_result

    level2_list = []
    fallback_unknown = False
    if hotel_inconsistent:
        level2_list.append('对比酒店不一致')
    if inout_inconsistent or roomtype_inconsistent or tongzhi_diff:
        level2_list.append('对比日期/房型/权益等不同')
    if is_free:
        level2_list.append('免费类权益lose')
    if is_jifen:
        level2_list.append('用户不认可积分优惠')
    if is_cuoluan:
        level2_list.append('错选乱选&未知原因')
    if is_unknown and not level2_list:
        level2_list.append('错选乱选&未知原因')

    if not level2_list:
        level2_list.append('错选乱选&未知原因')
        fallback_unknown = True

    level2_str = _join_values(level2_list)

    level3_list = []
    if inout_inconsistent:
        level3_list.append('对比日期不同')
    if hotel_inconsistent:
        level3_list.append('对比酒店不一致')
    if roomtype_inconsistent:
        level3_list.append('对比房型不同')
    if tongzhi_diff:
        level3_list.append('对比权益不同')

    if '错选乱选&未知原因' in level2_list:
        if is_cuoluan:
            level3_list.append('错选乱选')
        elif is_unknown or fallback_unknown:
            level3_list.append('未知原因')

    level3_str = _join_values(level3_list) if level3_list else ''

    return {'一级_AI': '用户误解', '二级_AI': level2_str, '三级_AI': level3_str,
            '四级_AI': '', '五级_AI': '', '归因备注': ''}


def attribute_price(row: dict, q_df: pd.DataFrame, c_df: pd.DataFrame) -> dict:
    case_id = str(row.get(FIELDS["case_no"], '') or '').strip()
    flow_no = str(row.get(FIELDS["order_no"], '') or '').strip()

    q_line = _safe_float(row.get(FIELDS["q_list_price"]))
    c_line = _safe_float(row.get(FIELDS["c_list_price"]))
    q_pay = _safe_float(row.get(FIELDS["q_pay_price"]))
    c_pay = _safe_float(row.get(FIELDS["c_pay_price"]))
    q_roomid = str(row.get(FIELDS["roomid_q"], '') or '').strip()
    c_roomid = str(row.get(FIELDS["roomid_c"], '') or '').strip()
    is_fenxiao = str(row.get(FIELDS["is_fenxiao"], '') or '').strip()
    is_louchu = str(row.get(FIELDS["is_louchu"], '') or '').strip()

    roomid_match = bool(q_roomid and c_roomid and q_roomid == c_roomid)
    lineprice_match = (q_line is not None and c_line is not None and abs(q_line - c_line) <= 0.1)

    if is_fenxiao == '过入住日期无法查询':
        return {'一级_AI': '无法归因', '二级_AI': '同CroomidQ未露出，且归因时入住日期，SPA无法查询',
                '三级_AI': '', '四级_AI': '', '五级_AI': '', '归因备注': ''}

    if (q_line is not None and c_line is not None and abs(q_line - c_line) > 0.1
            and is_fenxiao != '未分销'
            and '未露出' not in is_louchu):
        return {'一级_AI': '无法归因', '二级_AI': 'QCroomid相同，划线价不同',
                '三级_AI': '', '四级_AI': '', '五级_AI': '', '归因备注': ''}

    def q_sum(**filters):
        return sum_amounts_from_sheet(q_df, case_id, flow_no=flow_no, **filters)

    q_shangcu = q_sum(促销类型1级='商促')
    q_shangjuan = q_sum(促销类型1级='商券')
    q_pinjuan = q_sum(促销类型1级='平券')
    q_jiefen = q_sum(促销类型1级='积分')
    q_youxianghui = q_sum(促销类型2级='优享会')
    q_chuxing = q_sum(促销类型2级='出行身份')
    q_mendian = q_sum(促销类型2级='门店新客')
    q_xuesheng = q_sum(促销类型2级='学生身份')
    q_yoyo = q_sum(促销类型2级='yoyo卡')

    if q_df is not None and not q_df.empty and '是否十亿补贴' in q_df.columns:
        q_subset = _case_promos(q_df, case_id, flow_no)
        q_billion_mask = q_subset['是否十亿补贴'].astype(str).isin(['是', 'True', '1', 'true'])
        q_jiben = float(q_subset.loc[q_billion_mask, '促销金额'].apply(lambda x: _safe_float(x) or 0.0).sum()) if not q_subset.empty else 0.0
    else:
        q_jiben = 0.0

    def c_sum(**filters):
        return sum_amounts_from_sheet(c_df, case_id, flow_no=flow_no, **filters)

    c_shangcu = c_sum(促销类型1级='商促')
    c_shangjuan = c_sum(促销类型1级='商家券')
    c_pinjuan = c_sum(促销类型1级='平台券')
    c_jiefen = c_sum(促销类型1级='积分类')
    c_disan_butie = c_sum(促销类型1级='第三方补贴')

    def c_shenfenlei_sub(sub_keyword: str) -> float:
        if c_df is None or c_df.empty:
            return 0.0
        c_subset = _case_promos(c_df, case_id, flow_no)
        if c_subset.empty:
            return 0.0
        mask = c_subset['促销类型1级'].astype(str) == '身份类商促'
        sub_mask = mask & (
            c_subset['促销类型2级'].astype(str).str.contains(sub_keyword, na=False) |
            c_subset['促销类型3级'].astype(str).str.contains(sub_keyword, na=False)
        )
        return float(c_subset.loc[sub_mask, '促销金额'].apply(lambda x: _safe_float(x) or 0.0).sum())

    c_youxianghui = c_shenfenlei_sub('优享会')
    c_chuxing = c_shenfenlei_sub('出行身份')
    c_mendian = c_shenfenlei_sub('门店新客') + c_shenfenlei_sub('集团会员/首单')
    c_xuesheng = c_shenfenlei_sub('学生专享')
    c_jiben = c_sum(促销类型2级='十亿豪补')
    c_zhengfu = c_sum(促销类型2级='政府券')

    c_yoyo = 0.0
    if c_df is not None and not c_df.empty:
        c_subset = _case_promos(c_df, case_id, flow_no)
        if not c_subset.empty:
            c_yoyo = float(c_subset.loc[c_subset['促销名称'].astype(str).str.contains('yoyo卡', na=False), '促销金额']
                           .apply(lambda x: _safe_float(x) or 0.0).sum())

    q_quanqian = (q_pay or 0) - q_pinjuan if q_pay is not None else None
    c_quanqian = (c_pay or 0) - c_pinjuan - c_disan_butie if c_pay is not None else None

    is_homo = roomid_match and lineprice_match

    level2_list = []
    level3_list = []

    if not is_homo:
        if is_fenxiao == '未分销':
            return {'一级_AI': '价格lose', '二级_AI': '产品报价缺失', '三级_AI': '报价未接入',
                    '四级_AI': '', '五级_AI': '', '归因备注': ''}
        if '分销' in is_fenxiao and '未露出' in is_louchu:
            return {'一级_AI': '价格lose', '二级_AI': '产品报价缺失', '三级_AI': '报价未露出',
                    '四级_AI': '', '五级_AI': '', '归因备注': ''}

    if c_jiefen > q_jiefen:
        level2_list.append('积分补贴lose')
        level3_list.append('积分类lose')

    if c_disan_butie > 0:
        level2_list.append('支付补贴lose')
        level3_list.append('拿去花/信用购付款')

    promo_lose_level3 = []
    if c_jiben > q_jiben and q_quanqian is not None and c_quanqian is not None and q_quanqian > c_quanqian:
        promo_lose_level3.append('商促+平促混合包装lose')
    if c_shangcu > q_shangcu or c_shangjuan > q_shangjuan or c_yoyo > q_yoyo:
        promo_lose_level3.append('商促/券未接入')
    if c_pinjuan > q_pinjuan:
        promo_lose_level3.append('平台券lose')
    if (c_youxianghui > q_youxianghui or c_chuxing > q_chuxing or
            c_mendian > q_mendian or c_xuesheng > q_xuesheng):
        promo_lose_level3.append('身份类促销lose')
    if c_zhengfu > 0:
        promo_lose_level3.append('政府券lose')

    if promo_lose_level3:
        level2_list.append('促销lose')
        level3_list.extend(promo_lose_level3)

    level2_str = _join_values(level2_list) if level2_list else ''
    level3_str = _join_values(level3_list) if level3_list else ''

    level4_list = []
    l3_l4_paths = []
    if '平台券lose' in level3_str:
        lingjuan = str(row.get(FIELDS["lingjuan"], '') or '').strip()
        mapping = {
            '定价券lose': '定价券lose',
            '无对标券': '无对标券',
            '未领取券': '未领券',
            '未领券': '未领券',
            '已领取不能使用': '已领取不能使用',
            '领券无法使用': '已领取不能使用',
            '超过限制用券量': '超过限制用券量',
            '无法拆解': '无法拆解',
            '缺平台券领用记录': '无法拆解',
        }
        if lingjuan in mapping:
            val = mapping[lingjuan]
            level4_list.append(val)
            l3_l4_paths.append(('平台券lose', val))

    if '身份类促销lose' in level3_str:
        if c_youxianghui > q_youxianghui:
            level4_list.append('优享会')
            l3_l4_paths.append(('身份类促销lose', '优享会'))
        if c_chuxing > q_chuxing:
            level4_list.append('出行身份')
            l3_l4_paths.append(('身份类促销lose', '出行身份'))
        if c_mendian > q_mendian:
            level4_list.append('门店新客')
            l3_l4_paths.append(('身份类促销lose', '门店新客'))
        if c_xuesheng > q_xuesheng:
            level4_list.append('学生身份')
            l3_l4_paths.append(('身份类促销lose', '学生身份'))

    level4_str = _join_values(level4_list) if level4_list else ''

    return {'一级_AI': '价格lose', '二级_AI': level2_str, '三级_AI': level3_str,
            '四级_AI': level4_str, '五级_AI': '', '归因备注': '',
            '_l3_l4_paths': l3_l4_paths}


def attribute_price_tongcheng(row: dict, q_df: pd.DataFrame, c_df: pd.DataFrame) -> dict:
    case_id = str(row.get(FIELDS["case_no"], '') or '').strip()
    flow_no = str(row.get(FIELDS["order_no"], '') or '').strip()

    q_line = _safe_float(row.get(FIELDS["q_list_price"]))
    c_line = _safe_float(row.get(FIELDS["c_list_price"]))
    q_pay = _safe_float(row.get(FIELDS["q_pay_price"]))
    c_pay = _safe_float(row.get(FIELDS["c_pay_price"]))

    q_shangcu = sum_amounts_from_sheet(q_df, case_id, flow_no=flow_no, 促销类型1级='商促')
    q_shangjuan = sum_amounts_from_sheet(q_df, case_id, flow_no=flow_no, 促销类型1级='商券')
    q_pingcu = sum_amounts_from_sheet(q_df, case_id, flow_no=flow_no, 促销类型1级='平促')
    q_pinjuan = sum_amounts_from_sheet(q_df, case_id, flow_no=flow_no, 促销类型1级='平券')
    q_jiefen = sum_amounts_from_sheet(q_df, case_id, flow_no=flow_no, 促销类型1级='积分')
    q_shang_total = q_shangcu + q_shangjuan
    q_platform_total = q_pingcu + q_pinjuan

    e_shangcu = sum_amounts_from_sheet(c_df, case_id, flow_no=flow_no, 促销类型1级='商促')
    e_shangjuan = sum_amounts_from_sheet(c_df, case_id, flow_no=flow_no, 促销类型1级='商券')
    e_dingj = sum_amounts_from_sheet(c_df, case_id, flow_no=flow_no, 促销类型1级='平促')
    e_jiefen = sum_amounts_from_sheet(c_df, case_id, flow_no=flow_no, 促销类型1级='积分')
    e_fanxian = sum_amounts_from_sheet(c_df, case_id, flow_no=flow_no, 促销类型1级='返现')

    def is_e_common_platform_coupon(r):
        name = str(r.get('促销名称', ''))
        special = str(r.get('特殊类型', ''))
        amount = (_safe_float(r.get('促销金额')) or 0)
        return (str(r.get('促销类型1级', '')) == '平券'
                and special not in ('大额券', '黑鲸优惠')
                and name != '黑鲸优惠'
                and not (name == '可用券' and amount >= 50))

    e_pinjuan = _sum_promos(c_df, case_id, is_e_common_platform_coupon, flow_no=flow_no)
    e_daejuan = _sum_promos(
        c_df, case_id,
        lambda r: str(r.get('特殊类型', '')) == '大额券'
                  or (str(r.get('促销名称', '')) == '可用券' and (_safe_float(r.get('促销金额')) or 0) >= 50),
        flow_no=flow_no,
    )
    e_heijing = _sum_promos(
        c_df, case_id,
        lambda r: str(r.get('特殊类型', '')) == '黑鲸优惠' or str(r.get('促销名称', '')) == '黑鲸优惠',
        flow_no=flow_no,
    )
    e_shang_total = e_shangcu + e_shangjuan
    e_platform_total = e_dingj + e_pinjuan + e_daejuan + e_heijing

    if q_line is None or c_line is None:
        return {'一级_AI': '价格lose', '二级_AI': '', '三级_AI': '',
                '四级_AI': '', '五级_AI': '', '归因备注': '同程底价字段缺失',
                '_manual_reasons': ['价格类字段缺失:Q划线价或E划线价为空']}

    if abs(q_line - c_line) > 0.1:
        level3 = 'E划线价更低' if c_line < q_line else 'Q货源划线价更低'
        level4 = 'E含商促' if e_shang_total > 0 else 'E不含商促'
        return {'一级_AI': '价格lose', '二级_AI': '底价', '三级_AI': level3,
                '四级_AI': level4, '五级_AI': '', '归因备注': ''}

    level3_list = []
    level4_list = []
    l3_l4_paths = []
    manual_reasons = []

    if e_shang_total > q_shang_total:
        level3_list.append('商家补贴lose')
        names = _names_promos(
            c_df, case_id,
            lambda r: str(r.get('促销类型1级', '')) in ('商促', '商家券'),
            flow_no=flow_no,
        )
        level4_list.extend(names)
        l3_l4_paths.extend([('商家补贴lose', n) for n in names])

    if e_platform_total > q_platform_total:
        level3_list.append('平台补贴lose')
        e_platform_excl_heijing = e_dingj + e_pinjuan + e_daejuan
        l4_platform = []
        if e_platform_excl_heijing > q_platform_total:
            if e_dingj > q_pingcu:
                l4_platform.append('定价lose')
            if e_pinjuan > q_pinjuan:
                l4_platform.append('平券lose')
            if e_daejuan > 0:
                l4_platform.append('大额券lose')
            if e_heijing > 0:
                l4_platform.append('黑鲸优惠lose')
        else:
            l4_platform.append('黑鲸优惠lose')
        level4_list.extend(l4_platform)
        l3_l4_paths.extend([('平台补贴lose', v) for v in l4_platform])

    if e_jiefen > q_jiefen:
        level3_list.append('积分lose')
        multi_points = str(row.get(FIELDS["multi_points"], '') or '').strip()
        if multi_points in ('是', '多倍积分酒店') or '多倍' in multi_points and '非多倍' not in multi_points:
            level4_list.append('多倍积分酒店lose')
            l3_l4_paths.append(('积分lose', '多倍积分酒店lose'))
        elif multi_points in ('否', '非多倍积分酒店') or '非多倍' in multi_points:
            level4_list.append('非多倍积分酒店lose')
            l3_l4_paths.append(('积分lose', '非多倍积分酒店lose'))
        else:
            manual_reasons.append('价格类字段缺失:是否多倍积分酒店为空，积分lose四级待确认')

    if e_fanxian > 0:
        level3_list.append('返现lose')

    if not level3_list:
        manual_reasons.append(f'价格类但无法确定lose原因（Q支付价={q_pay}，E支付价={c_pay}）')

    return {'一级_AI': '价格lose',
            '二级_AI': '促销lose' if level3_list else '',
            '三级_AI': _join_values(level3_list),
            '四级_AI': _join_values(level4_list),
            '五级_AI': '',
            '归因备注': '',
            '_manual_reasons': manual_reasons,
            '_l3_l4_paths': l3_l4_paths}


# ═══════════════════════════════════════════
# 4. 主归因入口（单行）
# ═══════════════════════════════════════════

def _route_by_im(im_result: str) -> str:
    im = str(im_result or '').strip()
    if not im or im in ('空记录', ''):
        return ''
    if any(x in im for x in ['非C平台', '已下单', '民宿', '团购']):
        return '无法归因'
    if any(x in im for x in ['酒店不一致', '房型不一致', '入离不一致', '同质化权益',
                               '无相同浏览记录', '免费权益', '积分', '错选乱选', '未知原因', '未领券']):
        return '用户误解'
    if any(x in im for x in ['酒店缺失', '去哪儿房型缺失', '物理房型缺失', '房型缺失', '同质化房型缺失']):
        return '库存lose'
    if any(x in im for x in ['同质化lose', '价格高', '去哪儿价格高']):
        return '价格lose'
    return ''


HUMAN_REVIEW_CATS = ['金额校验', '字段缺失', '促销匹配', '归因不完整', '归因多重校验']
PROMO_AUDIT_KEEP_LEVEL1 = {'价格lose'}


def _is_promo_audit_reason(reason: str) -> bool:
    return '真实错误:V1' in reason or '真实错误:V2' in reason


def _classify_reason(reason: str) -> str:
    """将 manual_reason 字符串分类到5类人工校验类型之一"""
    if '真实错误:V1' in reason:
        return '金额校验'
    if '真实错误:V2' in reason:
        return '促销匹配'
    if '字段缺失' in reason:
        return '字段缺失'
    if '无法确定lose原因' in reason:
        return '归因不完整'
    if 'im=未知原因+cs不lose+Q含积分' in reason:
        return '归因不完整'
    if '路由冲突' in reason:
        return '归因多重校验'
    if reason.startswith('兜底验证'):
        return '归因多重校验'
    if reason.startswith('模糊场景') or reason.startswith('新场景'):
        return '归因不完整'
    return '归因不完整'


def attribute_row(row: dict, q_df: pd.DataFrame, c_df: pd.DataFrame,
                  price_rule_set: str = "xiecheng", platform: str = "携程",
                  promo_code_set: str = "xiecheng",
                  case_audit: dict = None, excel_row: int = None) -> dict:
    """对单行数据进行归因"""
    case_id = str(row.get(FIELDS["case_no"], '') or '').strip()
    case_audit = case_audit or {}
    # case编号为空时，用 row_N 与 validate.py 保持一致
    _audit_key = case_id if case_id else (f"row_{excel_row}" if excel_row else '')

    # ── 截图平台前置检查（最高优先级）──
    screenshot_platform = str(row.get(FIELDS["compare_platform"], '') or '').strip()
    if screenshot_platform and screenshot_platform not in ('', 'nan', 'None') and screenshot_platform not in VALID_PLATFORMS:
        return {'一级_AI': '无法归因', '二级_AI': '不可对比', '三级_AI': '非C平台',
                '四级_AI': '', '五级_AI': '', '归因备注': f'比价平台:{screenshot_platform}',
                '归因平台_AI': platform, 'price_rule_set_AI': price_rule_set,
                'promo_code_set_AI': promo_code_set}

    level1 = route_level1(row)

    v3_errors = validate_v3_consistency(level1, row)
    if v3_errors:
        return {'一级_AI': '校验失败', '二级_AI': 'V3错误',
                '三级_AI': '', '四级_AI': '', '五级_AI': '',
                '归因备注': '；'.join(v3_errors)}

    existing_note_value = row.get('归因备注', '')
    existing_note = '' if pd.isna(existing_note_value) else str(existing_note_value or '').strip()
    if existing_note in ('nan', 'None'):
        existing_note = ''

    if level1 == '异常':
        result = attribute_abnormal(row)
    elif level1 == '无法归因':
        result = attribute_unattributable(row)
    elif level1 == '库存lose':
        result = attribute_inventory(row)
    elif level1 == '用户误解':
        result = attribute_misconception(row)
    elif level1 == '价格lose':
        if price_rule_set == 'tongcheng':
            result = attribute_price_tongcheng(row, q_df, c_df)
        else:
            result = attribute_price(row, q_df, c_df)
    else:
        result = {'一级_AI': '异常', '二级_AI': '未知原因', '三级_AI': '',
                  '四级_AI': '', '五级_AI': '', '归因备注': ''}

    manual_reasons = result.pop('_manual_reasons', [])

    result['归因平台_AI'] = platform
    result['price_rule_set_AI'] = price_rule_set
    result['promo_code_set_AI'] = promo_code_set

    if existing_note and not result.get('归因备注'):
        result['归因备注'] = existing_note
    elif existing_note and result.get('归因备注'):
        result['归因备注'] = result['归因备注'] + '；' + existing_note

    # ── 路由冲突检查 ──
    im_result = str(row.get('im-归因结果-AI', '') or '').strip()
    kf = _text(row, FIELDS["kf_note"])
    r_im = _route_by_im(im_result)
    result['im一级推导'] = r_im or ''
    if r_im and r_im != level1:
        manual_reasons.append(f'路由冲突:kf={level1}/im={r_im}')

    # ── 兜底验证：cs 不lose 场景 ──
    cs_not_lose = '不lose' in kf
    if cs_not_lose:
        if '错选乱选' in im_result:
            manual_reasons.append('兜底验证:价格不lose+错选乱选')
        if '未知原因' in im_result:
            manual_reasons.append('兜底验证:价格不lose+未知原因')
        if '积分' in im_result:
            manual_reasons.append('兜底验证:价格不lose+积分')

    # ── 兜底验证：库存lose ──
    if result.get('一级_AI') == '库存lose':
        manual_reasons.append(f"兜底验证:库存lose-{result.get('二级_AI', '')}")

    # ── 兜底验证：im=未知原因+cs不lose+Q含积分 ──
    if '未知原因' in im_result and '不lose' in kf:
        q_case = _case_promos(q_df, case_id)
        if not q_case.empty and any(q_case['促销类型1级'].astype(str) == '积分'):
            manual_reasons.append('兜底验证:im=未知原因+cs不lose+Q含积分')

    # ── IM 归因特殊触发 ──
    if '无相同浏览记录' in im_result:
        manual_reasons.append('模糊场景:无相同浏览记录')
    im_explain = str(row.get('im-归因说明-AI', '') or '')
    if '新发现场景' in im_explain:
        manual_reasons.append(f'新场景:{im_result}')

    # ── 合并促销校验模块传递的 case_audit ──
    e_reasons = case_audit.get(_audit_key, [])
    if result.get('一级_AI') not in PROMO_AUDIT_KEEP_LEVEL1:
        e_reasons = [r for r in e_reasons if not _is_promo_audit_reason(r)]
    manual_reasons.extend(e_reasons)

    manual_categories = {cat: [] for cat in HUMAN_REVIEW_CATS}
    for r in manual_reasons:
        manual_categories[_classify_reason(r)].append(r)
    result['_manual_categories'] = manual_categories
    return result


# ═══════════════════════════════════════════
# 5. 主流程
# ═══════════════════════════════════════════

def _log(msg: str):
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}")


def _copy_promo_sheets(output_file: str, promo_file: str):
    """将促销分行表（Q/竞品）从促销数据文件复制到主数据文件。"""
    import openpyxl as _xl
    wb_promo = _xl.load_workbook(promo_file)
    wb_main = _xl.load_workbook(output_file)
    for sheet_name in ('Q促销分行表', '竞品促销分行表'):
        if sheet_name not in wb_promo.sheetnames:
            continue
        ws_src = wb_promo[sheet_name]
        if sheet_name in wb_main.sheetnames:
            del wb_main[sheet_name]
        ws_dst = wb_main.create_sheet(sheet_name)
        for row in ws_src.iter_rows():
            for cell in row:
                ws_dst.cell(cell.row, cell.column, cell.value)
    wb_main.save(output_file)


def _explode_last_level(df_keys: pd.DataFrame, group_keys: list) -> pd.DataFrame:
    """将最后一级列的多值字符串（"，"分隔）拆成独立行，每个场景各计1次。"""
    if not group_keys:
        return df_keys
    last_key = group_keys[-1]
    other_keys = group_keys[:-1]
    rows = []
    for _, row in df_keys.iterrows():
        raw = str(row[last_key]) if row[last_key] else ''
        parts = [v.strip() for v in raw.split('，') if v.strip()] or ['']
        for v in parts:
            new_row = {k: row[k] for k in other_keys}
            new_row[last_key] = v
            rows.append(new_row)
    return pd.DataFrame(rows, columns=group_keys) if rows else pd.DataFrame(columns=group_keys)


def _build_v4_paths_df(df_tmp: pd.DataFrame) -> pd.DataFrame:
    """V4 统计展开：利用 _l3_l4_paths 保证三→四级配对正确；无四级的路径以 '' 补位，保留所有行。"""
    level_cols = ['一级_AI', '二级_AI', '三级_AI', '四级_AI']
    has_paths = '_l3_l4_paths' in df_tmp.columns
    rows = []
    for _, r in df_tmp.iterrows():
        l1 = str(r.get('一级_AI', '') or '')
        l2 = str(r.get('二级_AI', '') or '')
        raw3 = str(r.get('三级_AI', '') or '')
        raw4 = str(r.get('四级_AI', '') or '')
        paths = r.get('_l3_l4_paths', []) if has_paths else []
        if isinstance(paths, list) and paths:
            # 有结构化路径：输出每个 (l3, l4) 对
            for (l3, l4) in paths:
                rows.append({'一级_AI': l1, '二级_AI': l2,
                             '三级_AI': str(l3) if l3 else '',
                             '四级_AI': str(l4) if l4 else ''})
            # 三级列中未出现在任何路径里的项（如返现lose）单独补一行，l4 为空
            l3_in_paths = {str(p[0]) for p in paths if p[0]}
            for l3 in [v.strip() for v in raw3.split('，') if v.strip()]:
                if l3 not in l3_in_paths:
                    rows.append({'一级_AI': l1, '二级_AI': l2, '三级_AI': l3, '四级_AI': ''})
        else:
            # 无结构化路径（非价格lose分支或旧数据）：逐项展开三级，l4 按旧逻辑处理
            parts3 = [v.strip() for v in raw3.split('，') if v.strip()] or ['']
            parts4 = [v.strip() for v in raw4.split('，') if v.strip()]
            if parts4:
                # 无法确定三四级配对，四级全挂在第一个三级下
                for p4 in parts4:
                    rows.append({'一级_AI': l1, '二级_AI': l2, '三级_AI': parts3[0], '四级_AI': p4})
                for l3 in parts3[1:]:
                    rows.append({'一级_AI': l1, '二级_AI': l2, '三级_AI': l3, '四级_AI': ''})
            else:
                # 无四级：每个三级项单独一行，l4 为空
                for l3 in parts3:
                    rows.append({'一级_AI': l1, '二级_AI': l2, '三级_AI': l3, '四级_AI': ''})
    return (pd.DataFrame(rows, columns=level_cols) if rows
            else pd.DataFrame(columns=level_cols))


def _generate_attribution_summary(output_file: str, result_df: pd.DataFrame):
    """生成归因结果汇总 sheet（4个递进版本，共用6列，上下排列）。
    多值字段（如三级="A，B"）按场景拆分，每场景单独计1，占比分母为原始行数。"""
    import openpyxl as _xl
    wb = _xl.load_workbook(output_file)
    sheet_name = '归因结果汇总'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    total = len(result_df)
    level_cols = ['一级_AI', '二级_AI', '三级_AI', '四级_AI']
    row_num = 1
    for version in range(1, 5):
        ws.cell(row_num, 1).value = f'▌ V{version}：{"→".join(str(i + 1) + "级" for i in range(version))}分布（场景拆分，占比/总案例数）'
        row_num += 1
        for c_idx, h in enumerate(level_cols + ['数量', '占比'], 1):
            ws.cell(row_num, c_idx).value = h
        row_num += 1
        group_keys = level_cols[:version]
        extra_cols = ['_l3_l4_paths'] if '_l3_l4_paths' in result_df.columns else []
        df_tmp = result_df[level_cols + extra_cols].copy()
        for col in level_cols:
            df_tmp[col] = df_tmp[col].fillna('').astype(str)
        if version == 4:
            df_exploded = _build_v4_paths_df(df_tmp)
        else:
            df_exploded = _explode_last_level(df_tmp[group_keys], group_keys)
        grouped = df_exploded.groupby(group_keys, dropna=False).size().reset_index(name='cnt')
        grouped = grouped.sort_values(group_keys).reset_index(drop=True)
        # 按一级分组：小计先行，明细后行
        l1_order, l1_groups = [], {}
        for _, r in grouped.iterrows():
            l1 = r['一级_AI']
            if l1 not in l1_groups:
                l1_order.append(l1)
                l1_groups[l1] = []
            l1_groups[l1].append(r)
        for l1 in l1_order:
            group_rows = l1_groups[l1]
            l1_cnt = sum(int(r['cnt']) for r in group_rows)
            if version > 1:
                ws.cell(row_num, 1).value = l1
                ws.cell(row_num, 2).value = '【小计】'
                for ci in range(3, 5):
                    ws.cell(row_num, ci).value = '/'
                ws.cell(row_num, 5).value = l1_cnt
                ws.cell(row_num, 6).value = f"{l1_cnt / total * 100:.1f}%" if total else '0.0%'
                row_num += 1
            for r in group_rows:
                for c_idx, k in enumerate(group_keys, 1):
                    ws.cell(row_num, c_idx).value = r[k]
                for c_idx in range(len(group_keys) + 1, 5):
                    ws.cell(row_num, c_idx).value = '/'
                ws.cell(row_num, 5).value = int(r['cnt'])
                ws.cell(row_num, 6).value = f"{r['cnt'] / total * 100:.1f}%" if total else '0.0%'
                row_num += 1
        row_num += 1
    wb.save(output_file)


def _generate_attribution_logic(output_file: str, result_df: pd.DataFrame):
    """生成归因逻辑 sheet（四级路径去重列出 + 归因规则列，供人工填写）。
    使用 _build_v4_paths_df 展开三级→四级配对，不统计数量。"""
    import openpyxl as _xl
    wb = _xl.load_workbook(output_file)
    sheet_name = '归因逻辑'
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    level_cols = ['一级_AI', '二级_AI', '三级_AI', '四级_AI']
    for c_idx, h in enumerate(level_cols + ['归因规则'], 1):
        ws.cell(1, c_idx).value = h
    extra_cols = ['_l3_l4_paths'] if '_l3_l4_paths' in result_df.columns else []
    df_tmp = result_df[level_cols + extra_cols].copy()
    for col in level_cols:
        df_tmp[col] = df_tmp[col].fillna('').astype(str)
    df_paths = _build_v4_paths_df(df_tmp)
    # 补入无四级的路径（三级分支）
    df_no_l4 = df_tmp[df_tmp['四级_AI'] == ''][level_cols].copy()
    df_no_l4 = df_no_l4.assign(
        三级_AI=df_no_l4['三级_AI'].apply(
            lambda v: [x.strip() for x in v.split('，') if x.strip()] or ['']
        )
    ).explode('三级_AI').reset_index(drop=True)
    unique_paths = (pd.concat([df_paths, df_no_l4], ignore_index=True)
                    .drop_duplicates(subset=level_cols)
                    .sort_values(level_cols)
                    .reset_index(drop=True))
    for row_num, (_, r) in enumerate(unique_paths.iterrows(), start=2):
        for c_idx, k in enumerate(level_cols, 1):
            ws.cell(row_num, c_idx).value = r[k]
        ws.cell(row_num, 5).value = ''
    wb.save(output_file)


def main():
    _log("attribute.py 启动")
    parser = argparse.ArgumentParser(description='归因脚本 — 五级归因树执行模块')
    parser.add_argument('--output-file', required=True, help='主数据文件路径（Sheet1=主数据）')
    parser.add_argument('--promo-file', required=False, default=None, help='促销数据文件路径（Sheet1=Q促销分行表, Sheet2=竞品促销分行表）')
    parser.add_argument('--state-file', required=False, default=None, help='Orchestrator state.json 路径')
    parser.add_argument('--platform', required=False, default=None, help='本次运行平台：携程/同程')
    parser.add_argument('--price-rule-set', required=False, default=None, help='价格规则集合：xiecheng/tongcheng')
    parser.add_argument('--promo-code-set', required=False, default=None, help='促销码表集合：xiecheng/tongcheng')
    args = parser.parse_args()

    if not os.path.exists(args.output_file):
        print(f"错误：文件不存在 - {args.output_file}", file=sys.stderr)
        sys.exit(1)

    state = load_full_state(args.state_file)
    state_input = state.get('input', {}) or {}

    run_platform = require_platform(args.platform or state_input.get('platform') or state_input.get('归因平台'))
    price_rule_set = normalize_rule_set(args.price_rule_set or state_input.get('price_rule_set'), run_platform)
    promo_code_set = normalize_rule_set(args.promo_code_set or state_input.get('promo_code_set'), run_platform)

    # 加载字段名映射（shared/field_map.json）
    loaded_fields = load_field_map(promo_code_set)
    if loaded_fields:
        FIELDS.update(loaded_fields)
        _log(f"字段映射已加载（promo_code_set={promo_code_set}，共 {len(loaded_fields)} 项）")
    else:
        _log("警告：field_map.json 未找到，使用默认字段名")

    # 确定促销数据文件路径
    promo_file = args.promo_file or state.get('promo_file')

    # 读取促销校验模块传递的 case_audit
    case_audit = {}
    if state and 'agents' in state and 'promo' in state['agents']:
        case_audit = state['agents']['promo'].get('case_audit', {})

    try:
        import openpyxl
    except ImportError:
        print("缺少依赖：pip3 install --break-system-packages openpyxl", file=sys.stderr)
        sys.exit(1)

    _log(f"运行平台: {run_platform} / price_rule_set={price_rule_set} / promo_code_set={promo_code_set}")
    _log(f"读取主数据文件: {args.output_file}")

    df = pd.read_excel(args.output_file, sheet_name="主数据", dtype=str)
    for col in [FIELDS["q_pay_price"], FIELDS["c_pay_price"], FIELDS["q_list_price"], FIELDS["c_list_price"]]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    _log(f"  → 主数据 {len(df)} 行")

    # 读取促销分行表（来自促销校验模块的独立文件）
    q_df = pd.DataFrame()
    c_df = pd.DataFrame()
    if promo_file and os.path.exists(promo_file):
        _log(f"读取促销数据文件: {promo_file}")
        try:
            q_df = pd.read_excel(promo_file, sheet_name="Q促销分行表", dtype=str)
            if '促销金额' in q_df.columns:
                q_df['促销金额'] = pd.to_numeric(q_df['促销金额'], errors='coerce').fillna(0.0)
            _log(f"  → Q促销分行表 {len(q_df)} 行")
        except Exception as e:
            print(f"  ⚠ Q促销分行表读取失败: {e}")
        try:
            # 优先读新命名 "竞品促销分行表"，兼容旧命名 "C促销分行表"
            for _sheet in ("竞品促销分行表", "C促销分行表"):
                try:
                    c_df = pd.read_excel(promo_file, sheet_name=_sheet, dtype=str)
                    break
                except Exception:
                    continue
            if c_df.empty:
                raise ValueError("未找到竞品促销分行表（尝试了 '竞品促销分行表' 和 'C促销分行表'）")
            if '促销金额' in c_df.columns:
                c_df['促销金额'] = pd.to_numeric(c_df['促销金额'], errors='coerce').fillna(0.0)
            _log(f"  → 竞品促销分行表 {len(c_df)} 行")
        except Exception as e:
            print(f"  ⚠ 竞品促销分行表读取失败: {e}")
    else:
        # 兜底：尝试从主数据文件本身读取内嵌的促销分行表
        _log("独立促销文件未找到，尝试从主数据文件读取内嵌促销分行表...")
        try:
            _wb_check = openpyxl.load_workbook(args.output_file, read_only=True)
            _sheets = _wb_check.sheetnames
            _wb_check.close()
            if 'Q促销分行表' in _sheets:
                q_df = pd.read_excel(args.output_file, sheet_name='Q促销分行表', dtype=str)
                if '促销金额' in q_df.columns:
                    q_df['促销金额'] = pd.to_numeric(q_df['促销金额'], errors='coerce').fillna(0.0)
                _log(f"  → Q促销分行表（来自主数据文件）{len(q_df)} 行")
            for _sheet in ('竞品促销分行表', 'C促销分行表'):
                if _sheet in _sheets:
                    c_df = pd.read_excel(args.output_file, sheet_name=_sheet, dtype=str)
                    if '促销金额' in c_df.columns:
                        c_df['促销金额'] = pd.to_numeric(c_df['促销金额'], errors='coerce').fillna(0.0)
                    _log(f"  → {_sheet}（来自主数据文件）{len(c_df)} 行")
                    break
            if q_df.empty and c_df.empty:
                print("  ⚠ 主数据文件中也未找到促销分行表，促销类归因金额将全按0处理")
        except Exception as e:
            print(f"  ⚠ 读取内嵌促销分行表失败: {e}，促销类归因金额将全按0处理")

    _log(f"开始执行归因, 共 {len(df)} 行")
    results = []
    manual_categories_list = []
    l3_l4_paths_list = []
    for i, (_, row) in enumerate(df.iterrows()):
        row_dict = row.to_dict()
        attr = attribute_row(row_dict, q_df, c_df, price_rule_set, run_platform, promo_code_set, case_audit, excel_row=i + 2)
        manual_categories_list.append(attr.pop('_manual_categories', {}))
        l3_l4_paths_list.append(attr.pop('_l3_l4_paths', []))
        results.append(attr)

    result_df = pd.DataFrame(results)
    result_df['_l3_l4_paths'] = l3_l4_paths_list

    target_cols = ['一级_AI', '二级_AI', '三级_AI', '四级_AI', '五级_AI', '归因备注',
                   '归因平台_AI', 'price_rule_set_AI', 'promo_code_set_AI', 'im一级推导']
    for col in target_cols:
        df[col] = result_df[col] if col in result_df.columns else ''

    # 写回主数据文件（保留单 sheet 结构）
    wb = openpyxl.load_workbook(args.output_file)
    ws = wb["主数据"]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h}

    for col_name in target_cols:
        if col_name not in col_map:
            new_idx = ws.max_column + 1
            ws.cell(1, new_idx).value = col_name
            col_map[col_name] = new_idx

    # 人工校验列（分类结构）
    _manual_col_names = ['是否需要人工校验', '人工校验分类'] + [f'{c}_详情' for c in HUMAN_REVIEW_CATS]
    for col_name in _manual_col_names:
        if col_name not in col_map:
            new_idx = ws.max_column + 1
            ws.cell(1, new_idx).value = col_name
            col_map[col_name] = new_idx

    manual_flag_col = col_map['是否需要人工校验']
    manual_class_col = col_map['人工校验分类']
    manual_detail_cols = {cat: col_map[f'{cat}_详情'] for cat in HUMAN_REVIEW_CATS}

    for i, (_, result_row) in enumerate(result_df.iterrows()):
        excel_row = i + 2
        for col_name in target_cols:
            col_idx = col_map[col_name]
            ws.cell(excel_row, col_idx).value = result_row.get(col_name, '') or ''

        cats = manual_categories_list[i]
        active_cats = [cat for cat in HUMAN_REVIEW_CATS if cats.get(cat)]
        ws.cell(excel_row, manual_flag_col).value = ''
        ws.cell(excel_row, manual_class_col).value = ''
        for cat in HUMAN_REVIEW_CATS:
            ws.cell(excel_row, manual_detail_cols[cat]).value = ''
        if active_cats:
            ws.cell(excel_row, manual_flag_col).value = '是'
            ws.cell(excel_row, manual_class_col).value = '、'.join(active_cats)
            for cat in active_cats:
                ws.cell(excel_row, manual_detail_cols[cat]).value = '；'.join(cats[cat])

    wb.save(args.output_file)
    _log(f"归因结果已写回主数据: {args.output_file}")

    if promo_file and os.path.exists(promo_file):
        _log("将促销分行表写入主数据文件...")
        _copy_promo_sheets(args.output_file, promo_file)
        os.remove(promo_file)
        _log(f"促销分行表写入完成，独立促销文件已删除: {promo_file}")

    _log("生成归因结果汇总...")
    _generate_attribution_summary(args.output_file, result_df)

    _log("生成归因逻辑...")
    _generate_attribution_logic(args.output_file, result_df)

    _log(f"全部完成: {args.output_file}")

    print('\n── 一级_AI 分布 ──')
    print(df['一级_AI'].value_counts().to_string())
    print('\n── 二级_AI 分布 ──')
    print(df['二级_AI'].value_counts().to_string())
    print('\n── 三级_AI 分布 ──')
    print(df['三级_AI'].value_counts().to_string())

    manual_count = sum(1 for cats in manual_categories_list if any(cats.values()))
    print(f'\n人工校验标记：{manual_count} 条')
    check_items = df[df.get('归因备注', pd.Series(dtype=str)).str.contains('错误|未匹配|异常', na=False)]
    print(f'需关注: {len(check_items)} 条（含校验错误）')


if __name__ == '__main__':
    main()
