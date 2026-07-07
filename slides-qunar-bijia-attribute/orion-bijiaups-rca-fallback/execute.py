#!/usr/bin/env python3
"""
execute.py — fallback 模块 Excel 读写工具

Commands:
  read-case   从原始 Excel 读取指定 case 的字段值，输出 JSON 到 stdout
  write-promo 更新促销分行表的兜底结果列（支持 --side q|c|both）
  check-v2    对指定促销名称列表做码表匹配，返回哪些已匹配/未匹配
"""

import argparse
import json
import sys
import os
from datetime import datetime

import openpyxl

# 导入 validate.py 的码表加载和 enrich 函数（用于 check-v2）
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'orion-bijiaups-rca-promo'))
try:
    from validate import (
        normalize_promo_name,
        enrich_q_promos,
        enrich_c_promos,
        load_code_table_q,
        load_code_table_c,
    )
    _VALIDATE_AVAILABLE = True
except ImportError:
    _VALIDATE_AVAILABLE = False

CODE_TABLE_PATH_TONGCHENG = os.path.join(
    os.path.dirname(__file__), '..', 'orion-bijiaups-rca-promo', 'code_tables', 'E_QE.xlsx'
)
CODE_TABLE_PATH_Q_DEFAULT = os.path.join(
    os.path.dirname(__file__), '..', 'orion-bijiaups-rca-promo', 'code_tables', 'C_Q.xlsx'
)
CODE_TABLE_PATH_C_DEFAULT = os.path.join(
    os.path.dirname(__file__), '..', 'orion-bijiaups-rca-promo', 'code_tables', 'C_C.xlsx'
)


def _log(msg: str):
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", file=sys.stderr)


def _safe(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    return str(val).strip()


# ─── read-case ────────────────────────────────────────────────────────────────

def cmd_read_case(input_file: str, case_id: str, fields: list) -> dict:
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb.active
    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    case_col = headers.index('case编号') if 'case编号' in headers else None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_case_id = str(row[case_col] or '').strip() if case_col is not None else ''
        row_key = row_case_id if row_case_id else f"row_{row_idx}"

        if row_key == case_id or (row_case_id and row_case_id == case_id):
            return {
                field: _safe(row[headers.index(field)]) if field in headers else None
                for field in fields
            }

    return {}


# ─── check-v2 ─────────────────────────────────────────────────────────────────

def cmd_check_v2(promo_names: list, side: str, promo_code_set: str) -> dict:
    """
    对给定促销名称列表做码表匹配，返回匹配结果。
    side: q | c
    返回: {"all_matched": bool, "matched": [...], "unmatched": [...]}
    """
    if not _VALIDATE_AVAILABLE:
        return {"all_matched": False, "matched": [], "unmatched": promo_names,
                "error": "validate.py not importable"}

    # 加载码表
    if promo_code_set == "tongcheng":
        code_q = load_code_table_q(CODE_TABLE_PATH_TONGCHENG, sheet_name="Q促销码表")
        code_c = load_code_table_c(CODE_TABLE_PATH_TONGCHENG, sheet_name="E促销码表")
    else:
        code_q = load_code_table_q(CODE_TABLE_PATH_Q_DEFAULT)
        code_c = load_code_table_c(CODE_TABLE_PATH_C_DEFAULT)

    # 构造 promo_dicts
    promo_dicts = [
        {"raw_name": n, "name_norm": normalize_promo_name(n), "amount": 0.0}
        for n in promo_names if n
    ]

    # enrich
    if side == "q":
        enriched = enrich_q_promos(promo_dicts, code_q)
    else:
        enriched = enrich_c_promos(promo_dicts, code_c, promo_code_set)

    matched   = [p["raw_name"] for p in enriched if p.get("matched")]
    unmatched = [p["raw_name"] for p in enriched if not p.get("matched")]

    return {
        "all_matched": len(unmatched) == 0 and len(enriched) > 0,
        "matched": matched,
        "unmatched": unmatched,
    }


# ─── write-promo ──────────────────────────────────────────────────────────────

def cmd_write_promo(promo_file: str, case_id: str, tier: str,
                    optimization: str = '',
                    q_amounts: list = None, c_amounts: list = None,
                    v1_resolved: bool = False,
                    v2_resolved: bool = False,
                    side: str = "both"):
    """
    更新促销分行表兜底结果。

    side: "q" | "c" | "both"
      - "q"   → 只写 Q促销分行表
      - "c"   → 只写 竞品促销分行表
      - "both"→ 两张表都写（向后兼容）

    V1/V2 独立处理：
      v1_resolved=True → V1校验列写"通过{base_tier}已解决"（原有错误信息才改）
      v2_resolved=True → V2校验列"未匹配"行改为"通过{base_tier}已解决"
      未解决的列保持原样不动
    """
    wb = openpyxl.load_workbook(promo_file)

    # base_tier：去掉后缀，用于标注 V1/V2 列（如"兜底1-V1待处理"→"兜底1"）
    base_tier = tier.split("-")[0]

    sheets_to_write = []
    if side in ("q", "both"):
        sheets_to_write.append(("Q促销分行表", q_amounts))
    if side in ("c", "both"):
        sheets_to_write.append(("竞品促销分行表", c_amounts))

    for sheet_name, amounts in sheets_to_write:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

        def ensure_col(name):
            if name not in headers:
                idx = ws.max_column + 1
                ws.cell(1, idx).value = name
                headers.append(name)
            return headers.index(name) + 1

        tier_col   = ensure_col("兜底方式")
        opt_col    = ensure_col("兜底优化建议")
        case_col   = headers.index("case编号") + 1 if "case编号" in headers else None
        amount_col = headers.index("促销金额") + 1 if "促销金额" in headers else None
        v1_col     = headers.index("V1校验") + 1   if "V1校验" in headers else None
        v2_col     = headers.index("V2校验") + 1   if "V2校验" in headers else None

        if case_col is None:
            continue

        case_rows = [
            row_cells for row_cells in ws.iter_rows(min_row=2)
            if str(row_cells[case_col - 1].value or '').strip() == str(case_id).strip()
        ]

        for list_idx, row_cells in enumerate(case_rows):
            row_num = row_cells[0].row

            # 兜底方式（所有情况都写）
            ws.cell(row_num, tier_col).value = tier

            # 优化建议（有内容才写）
            if optimization:
                ws.cell(row_num, opt_col).value = optimization

            # 促销金额（V1解决时按行覆写）
            if amounts and amount_col and list_idx < len(amounts):
                corrected = amounts[list_idx]
                if corrected is not None:
                    ws.cell(row_num, amount_col).value = corrected

            # V1校验列：解决了写"通过兜底x已解决"，未解决保持原样
            if v1_resolved and v1_col:
                orig_v1 = ws.cell(row_num, v1_col).value
                if orig_v1:  # 有原始错误才改
                    ws.cell(row_num, v1_col).value = f"通过{base_tier}已解决"

            # V2校验列：解决了把"未匹配"改为"通过兜底x已解决"，未解决保持"未匹配"
            if v2_resolved and v2_col:
                orig_v2 = ws.cell(row_num, v2_col).value
                if orig_v2 == "未匹配":
                    ws.cell(row_num, v2_col).value = f"通过{base_tier}已解决"

    wb.save(promo_file)
    _log(f"write-promo 完成: case_id={case_id}, tier={tier}, side={side}, v1={v1_resolved}, v2={v2_resolved}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="fallback 模块 Excel 读写工具")
    sub = parser.add_subparsers(dest="command")

    # read-case
    p_read = sub.add_parser("read-case", help="读取指定 case 的字段值")
    p_read.add_argument("--input-file", required=True)
    p_read.add_argument("--case-id", required=True)
    p_read.add_argument("--fields", required=True, help="逗号分隔的字段名")

    # check-v2
    p_chk = sub.add_parser("check-v2", help="对促销名称列表做码表匹配")
    p_chk.add_argument("--promo-names", required=True, help="JSON数组，促销名称列表")
    p_chk.add_argument("--side", required=True, choices=["q", "c"])
    p_chk.add_argument("--promo-code-set", required=True, help="tongcheng | xiecheng")

    # write-promo
    p_write = sub.add_parser("write-promo", help="更新促销分行表兜底结果列")
    p_write.add_argument("--promo-file", required=True)
    p_write.add_argument("--case-id", required=True)
    p_write.add_argument("--tier", required=True,
                         help="兜底1|兜底1-V1待处理|兜底1-V2待码表更新|兜底2|兜底2-V1待处理|兜底2-V2待码表更新|需人工介入")
    p_write.add_argument("--optimization", default="")
    p_write.add_argument("--q-amounts", default=None, help="JSON数组，Q侧各行修正金额")
    p_write.add_argument("--c-amounts", default=None, help="JSON数组，竞品侧各行修正金额")
    p_write.add_argument("--v1-resolved", action="store_true", help="V1金额已平衡")
    p_write.add_argument("--v2-resolved", action="store_true", help="V2类型已匹配码表")
    p_write.add_argument("--side", default="both", choices=["q", "c", "both"],
                         help="只写指定侧分行表（默认both向后兼容）")

    args = parser.parse_args()

    if args.command == "read-case":
        fields = [f.strip() for f in args.fields.split(",")]
        result = cmd_read_case(args.input_file, args.case_id, fields)
        print(json.dumps(result, ensure_ascii=False))

    elif args.command == "check-v2":
        promo_names = json.loads(args.promo_names)
        result = cmd_check_v2(promo_names, args.side, args.promo_code_set)
        print(json.dumps(result, ensure_ascii=False))

    elif args.command == "write-promo":
        q_amounts = json.loads(args.q_amounts) if args.q_amounts else None
        c_amounts = json.loads(args.c_amounts) if args.c_amounts else None
        cmd_write_promo(
            args.promo_file, args.case_id, args.tier,
            optimization=args.optimization,
            q_amounts=q_amounts, c_amounts=c_amounts,
            v1_resolved=args.v1_resolved,
            v2_resolved=args.v2_resolved,
            side=args.side,
        )

    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
