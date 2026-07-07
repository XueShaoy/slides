#!/usr/bin/env python3
"""
IM归因 Excel I/O 工具
Commands:
  init  --input <file.xlsx> [--output <out.xlsx>]   # 创建主数据输出文件（Sheet1=原始数据）
  read  --input <file.xlsx>                          # 读取IM对话数据，输出JSON（从原始输入文件读）
  write --output <out.xlsx> --results <results.json> # 写回im-归因结果-AI/im-归因说明-AI 到Sheet1
"""

# ============================================================
# ⚠️ 修改联动声明
# 本脚本与 io.md（同目录）强绑定。
# 改接口行为：先更新 io.md → 再修改本脚本
# 改代码：先确认 io.md 中的接口描述 → 修改后同步更新 io.md
# 两者不同步 = 接口文档与代码行为不一致
# ============================================================

import sys
import os
import json
import shutil
import argparse
from pathlib import Path
from datetime import datetime

# ── 字段名映射（运行时由 shared/field_map.json 覆盖） ──
# 修改字段名请同步更新 shared/field_map.json 和 shared/field_definitions.md
_FIELD_MAP_PATH = Path(__file__).parent / "field_map.json"
FIELDS: dict = {
    "im_text":  "客服与用户沟通记录",
    "kf_note":  "客服判断场景",
    "case_no":  "case编号",
    "order_no": "工单号",
}
try:
    _raw = json.loads(_FIELD_MAP_PATH.read_text(encoding="utf-8"))
    FIELDS.update(_raw.get("common", {}))
except Exception:
    pass


def check_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("缺少依赖：pip3 install --break-system-packages openpyxl", file=sys.stderr)
        sys.exit(1)


def make_output_path(input_path):
    """生成输出文件路径：原目录/原文件名_AI主数据_YYYYmmdd_HHMM.xlsx"""
    dir_name = os.path.dirname(os.path.abspath(input_path))
    base = os.path.splitext(os.path.basename(input_path))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    return os.path.join(dir_name, f"{base}_AI主数据_{timestamp}.xlsx")


def _log(msg: str):
    """结构化日志：时间戳 + 消息"""
    from datetime import datetime
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", file=sys.stderr)


def init_output_file(input_path, output_path=None):
    """
    创建主数据输出文件：
      Sheet1（主数据）：原始数据完整复制
    促销数据由促销校验模块单独创建，不在此处生成。
    """
    _log(f"init_output_file 开始, input={input_path}")
    openpyxl = check_openpyxl()

    if output_path is None:
        output_path = make_output_path(input_path)

    shutil.copy2(input_path, output_path)

    wb = openpyxl.load_workbook(output_path)

    # 确保"主数据" sheet 存在且名称正确
    if "主数据" not in wb.sheetnames:
        target = None
        for ws in wb.worksheets:
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            if "case编号" in headers and ws.max_row > 5:
                target = ws
                break
        if target is None:
            target = wb.worksheets[0]
        target.title = "主数据"

    # 清理所有多余 sheet（只保留"主数据"）
    for name in [s for s in wb.sheetnames if s != "主数据"]:
        del wb[name]

    wb.save(output_path)
    _log(f"init_output_file 完成, 输出={output_path}")
    print(f"OUTPUT_PATH:{output_path}")
    print(f"完成！主数据文件已创建：{output_path}")


def read_xlsx(input_path):
    """从输入文件读取IM对话数据，输出JSON数组"""
    openpyxl = check_openpyxl()
    wb = openpyxl.load_workbook(input_path, data_only=True)

    if "主数据" in wb.sheetnames:
        ws = wb["主数据"]
    else:
        ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    im_col = None
    case_col = None
    order_col = None
    cs_judgment_col = None

    for i, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h)
        if h_str == FIELDS["im_text"]:
            im_col = i + 1
        if h_str == FIELDS["case_no"]:
            case_col = i + 1
        if h_str == FIELDS["order_no"]:
            order_col = i + 1
        if h_str == FIELDS["kf_note"]:
            cs_judgment_col = i + 1

    if not im_col:
        print(f"错误：未找到IM对话列（列名必须精确为'{FIELDS['im_text']}'）。当前列：{headers}", file=sys.stderr)
        sys.exit(1)

    records = []
    for row in range(2, ws.max_row + 1):
        im_text = ws.cell(row, im_col).value
        case_id = ws.cell(row, case_col).value if case_col else None
        flow_no = ws.cell(row, order_col).value if order_col else None
        cs_judgment = ws.cell(row, cs_judgment_col).value if cs_judgment_col else None

        records.append({
            "row_num": row,
            "case_id": str(case_id).strip() if case_id else f"row_{row}",
            "flow_no": str(flow_no) if flow_no else "",
            "im_text": str(im_text).strip() if im_text else "",
            "cs_judgment": str(cs_judgment).strip() if cs_judgment else ""
        })

    print(json.dumps(records, ensure_ascii=False, indent=2))


def write_xlsx(output_path, results_data):
    """
    将im归因结果写入主数据文件的Sheet1（主数据）。
    只写 im-归因结果-AI 和 im-归因说明-AI 两列，人工校验列由归因模块统一写入。
    结果JSON格式：[{"row_num": 2, "case_id": "...", "short_labels": "...", "explanation": "..."}]
    """
    _log(f"write_xlsx 开始, output={output_path}, 结果数={len(results_data)}")
    openpyxl = check_openpyxl()

    wb = openpyxl.load_workbook(output_path)

    if "主数据" in wb.sheetnames:
        ws = wb["主数据"]
    else:
        ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    label_col = None
    explain_col = None
    for i, h in enumerate(headers):
        if h == "im-归因结果-AI":
            label_col = i + 1
        if h == "im-归因说明-AI":
            explain_col = i + 1

    if label_col is None:
        label_col = ws.max_column + 1
        ws.cell(1, label_col).value = "im-归因结果-AI"
    if explain_col is None:
        explain_col = max(label_col, ws.max_column) + 1
        ws.cell(1, explain_col).value = "im-归因说明-AI"

    results_map = {r["row_num"]: r for r in results_data}

    written = 0
    for row in range(2, ws.max_row + 1):
        if row in results_map:
            r = results_map[row]
            ws.cell(row, label_col).value = r.get("short_labels", "未知原因")
            ws.cell(row, explain_col).value = r.get("explanation", "")
            written += 1

    wb.save(output_path)
    _log(f"write_xlsx 完成, 写入 {written} 条")
    print(f"完成！写入 {written} 条。结果文件：{output_path}")


def main():
    parser = argparse.ArgumentParser(description="IM归因 Excel I/O 工具")
    subparsers = parser.add_subparsers(dest="command")

    init_p = subparsers.add_parser("init", help="创建主数据输出文件")
    init_p.add_argument("--input", required=True, help="原始输入Excel文件路径")
    init_p.add_argument("--output", required=False, default=None, help="输出文件路径（不填则自动生成）")

    read_p = subparsers.add_parser("read", help="读取IM对话数据，输出JSON")
    read_p.add_argument("--input", required=True, help="输入Excel文件路径")

    write_p = subparsers.add_parser("write", help="写回im归因结果到主数据文件")
    write_p.add_argument("--output", required=True, help="主数据文件路径")
    write_p.add_argument("--results", required=True, help="结果JSON文件路径")

    args = parser.parse_args()

    if args.command == "init":
        if not os.path.exists(args.input):
            print(f"错误：文件不存在 - {args.input}", file=sys.stderr)
            sys.exit(1)
        init_output_file(args.input, args.output)

    elif args.command == "read":
        if not os.path.exists(args.input):
            print(f"错误：文件不存在 - {args.input}", file=sys.stderr)
            sys.exit(1)
        read_xlsx(args.input)

    elif args.command == "write":
        for p in [args.output, args.results]:
            if not os.path.exists(p):
                print(f"错误：文件不存在 - {p}", file=sys.stderr)
                sys.exit(1)
        with open(args.results, "r", encoding="utf-8") as f:
            results_data = json.load(f)
        write_xlsx(args.output, results_data)

    else:
        parser.print_help()


if __name__ == "__main__":
    main()
