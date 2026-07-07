#!/usr/bin/env python3
"""
促销校验模块
功能：
  1. 从原始输入 Excel 读取价格和促销数据（不依赖 IM归因模块的输出文件）
  2. V1 金额校验：|划线价 - (促销总金额 + 支付价)| ≤ 1.0
  3. V2 促销归类校验：所有促销名称必须在码表中精确命中
  4. 比价平台校验：不在[携程/美团/飞猪/同程]中 → 标记不可比价
  5. 解析促销明细，匹配码表类型，写入独立促销数据文件（Sheet1=Q促销分行表, Sheet2=竞品促销分行表）
  6. 将 case_audit（每 case 的人工校验原因）和 promo_file 路径写入 state.json
  7. 输出校验结果JSON到标准输出

字段名通过 field_map.json 配置，详见 shared/field_definitions.md。

用法：
  python3 validate.py --state-file <state.json>
  # 也可显式传入参数覆盖 state.json 中的值：
  python3 validate.py --input-file <raw.xlsx> --promo-file <promo.xlsx> --state-file <state.json>
  # 兼容旧调用：也可显式传 --code-q / --code-c 覆盖内置码表
"""

# ============================================================
# ⚠️ 修改联动声明
# 本脚本与 promo_logic.md（同目录）强绑定。
# 改规则：先更新 promo_logic.md → 再修改本脚本
# 改代码：先确认 promo_logic.md 中的规则 → 修改后同步更新 promo_logic.md
# 两者不同步 = 归因结果不可信
# ============================================================

import argparse
import json
import re
import sys
import os
from datetime import datetime
from pathlib import Path


VALID_PLATFORMS = {"携程", "美团", "飞猪", "同程"}
RULE_ROOT = Path("/Users/zhangwang/.claude/agents/orion-bijiaups-rca-shared")
DEFAULT_CODE_Q = "/Users/zhangwang/.claude/agents/orion-bijiaups-rca-promo/code_tables/C_Q.xlsx"
DEFAULT_CODE_C = "/Users/zhangwang/.claude/agents/orion-bijiaups-rca-promo/code_tables/C_C.xlsx"

PLATFORM_ALIASES = {
    "xiecheng": "携程",
    "ctrip": "携程",
    "携程": "携程",
    "tongcheng": "同程",
    "tc": "同程",
    "同程": "同程",
}

PROMO_CODE_SET_BY_PLATFORM = {
    "携程": "xiecheng",
    "同程": "tongcheng",
}

# ── 字段名映射（运行时由 field_map.json 覆盖） ──
# 修改字段名请同步更新 shared/field_map.json 和 shared/field_definitions.md
FIELDS: dict = {
    "kf_note":          "客服判断场景",
    "compare_platform": "比价平台",
    "q_list_price":     "Q划线价——同质化房型",
    "q_pay_price":      "Q到手价——同质化房型",
    "c_list_price":     "划线价（竞品）",
    "c_pay_price":      "到手价（竞品）",
    "q_promo_detail":   "Q优惠明细",
    "c_promo_detail":   "促销明细（竞品）",
    "case_no":          "case编号",
    "order_no":         "工单号",
}


def load_field_map(promo_code_set: str) -> dict:
    """从 field_map.json 加载字段名映射（common + 平台专用合并）。"""
    map_path = RULE_ROOT / "field_map.json"
    if not map_path.exists():
        return {}
    with open(map_path, encoding="utf-8") as f:
        data = json.load(f)
    common = data.get("common", {})
    platform_specific = data.get(promo_code_set, {})
    return {**common, **platform_specific}


# 去除噪音符号（保留 -）
_NOISE_RE = re.compile(r'[_。，、；：．\.\s]')
# 促销名称中"返￥XX"尾缀的金额部分（normalize时剥掉数字，保留"返"）
_FAN_AMOUNT_RE = re.compile(r'返[¥￥]([\d,.]+)\s*$')

# 匹配格式（兼容两种格式）：
#   Q侧（携程）：促销名称 -¥金额 / 促销名称——¥金额
#   竞品侧（同程）：促销名称：金额 / 促销名称:金额（全角/半角冒号，无¥符号）
_PROMO_RE = re.compile(r'^(.+?)(?:\s*[-—–]+\s*[¥￥]?\s*|[：:]\s*)([\d,.]+)\s*$')
# "促销名 返￥金额" 格式（如：惊喜特惠券 返￥50）
_PROMO_RE_FAN = re.compile(r'^(.+?)\s*返[¥￥]([\d,.]+)\s*$')

# 客服判断场景字段的不可对比关键词（命中则跳过V1/V2校验和促销分行）
_KF_NONCOMPARABLE = [
    '民宿', '团购', '预售', '已下单',
    '不可对比平台', '非可对比平台', '非可比价平台',
    '物理房型', '酒店缺失', '无该酒店', '没有该酒店',
    '更好更贵', '更差更贵',
]


def check_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("缺少依赖：pip3 install --break-system-packages openpyxl", file=sys.stderr)
        sys.exit(1)


# ─── 促销名称标准化 ───

def normalize_promo_name(name: str) -> str:
    # 先剥掉"返￥金额"中的金额数字（保留"返"），使"惊喜特惠券 返￥50"→"惊喜特惠券 返"与码表对齐
    s = _FAN_AMOUNT_RE.sub('返', str(name))
    s = _NOISE_RE.sub('', s)
    return s.strip()


# ─── 码表加载 ───

def normalize_platform(value: str | None) -> str:
    key = str(value or '').strip()
    return PLATFORM_ALIASES.get(key, key)


def require_platform(value: str | None) -> str:
    platform = normalize_platform(value)
    if not platform:
        print("错误：缺少归因平台。当前单次运行必须明确指定平台：携程 / 同程", file=sys.stderr)
        sys.exit(1)
    if platform not in PROMO_CODE_SET_BY_PLATFORM:
        print(f"错误：暂不支持的归因平台 - {platform}。当前支持：携程 / 同程", file=sys.stderr)
        sys.exit(1)
    return platform


def normalize_rule_set(value: str | None, platform: str) -> str:
    key = str(value or '').strip()
    if key in ("xiecheng", "tongcheng"):
        return key
    return PROMO_CODE_SET_BY_PLATFORM.get(platform, "xiecheng")


def load_state_options(state_file: str | None) -> dict:
    if not state_file:
        return {}
    if not os.path.exists(state_file):
        print(f"错误：state.json不存在 - {state_file}", file=sys.stderr)
        sys.exit(1)
    with open(state_file, "r", encoding="utf-8") as f:
        state = json.load(f)
    return state.get("input", {}) or {}


def load_full_state(state_file: str | None) -> dict:
    if not state_file or not os.path.exists(state_file):
        return {}
    with open(state_file, "r", encoding="utf-8") as f:
        return json.load(f)


def resolve_code_table_paths(args, platform: str, promo_code_set: str) -> tuple[str, str, str | None, str | None]:
    if args.code_q and args.code_c:
        return args.code_q, args.code_c, None, None
    if promo_code_set == "tongcheng":
        path = "/Users/zhangwang/.claude/agents/orion-bijiaups-rca-promo/code_tables/E_QE.xlsx"
        return path, path, "Q促销码表", "E促销码表"
    return args.code_q or DEFAULT_CODE_Q, args.code_c or DEFAULT_CODE_C, None, None


def load_code_table_q(path: str, sheet_name: str | None = None) -> dict:
    openpyxl = check_openpyxl()
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else '' for h in rows[0]]

    name_col = headers.index('促销名称')
    type1_col = headers.index('促销类型1级')
    type2_col = headers.index('促销类型2级') if '促销类型2级' in headers else None
    type3_col = headers.index('促销类型3级') if '促销类型3级' in headers else None
    billion_col = headers.index('十亿补贴') if '十亿补贴' in headers else None
    special_col = headers.index('特殊类型') if '特殊类型' in headers else None

    table = {}
    for row in rows[1:]:
        if not row[name_col]:
            continue
        raw_name = str(row[name_col]).strip()
        key = normalize_promo_name(raw_name)
        type1 = str(row[type1_col]).strip() if row[type1_col] else ''
        type2 = str(row[type2_col]).strip() if type2_col is not None and row[type2_col] else ''
        type3 = str(row[type3_col]).strip() if type3_col is not None and row[type3_col] else ''
        is_billion = bool(row[billion_col]) if billion_col is not None else False
        special_type = str(row[special_col]).strip() if special_col is not None and row[special_col] else ''
        table[key] = {
            'type1': type1, 'type2': type2, 'type3': type3,
            'is_billion': is_billion, 'special_type': special_type, 'raw_name': raw_name,
        }
    return table


def load_code_table_c(path: str, sheet_name: str | None = None) -> dict:
    openpyxl = check_openpyxl()
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else '' for h in rows[0]]

    name_col = None
    for i, h in enumerate(headers):
        if h and ('促销' in h or '权益' in h) and '名称' in h:
            name_col = i
            break
    if name_col is None:
        name_col = 0

    type1_raw_col = headers.index('促销类型1级') if '促销类型1级' in headers else None
    type2_col = headers.index('促销类型2级') if '促销类型2级' in headers else None
    type3_col = headers.index('促销类型3级') if '促销类型3级' in headers else None
    special_col = headers.index('特殊类型') if '特殊类型' in headers else None

    code_col = None
    for i, h in enumerate(headers):
        if h and '码表用' in h and '(1)' not in h:
            code_col = i
            break

    table = {}
    for row in rows[1:]:
        if not row[name_col]:
            continue
        raw_name = str(row[name_col]).strip()
        key = normalize_promo_name(raw_name)
        type1_raw = str(row[type1_raw_col]).strip() if type1_raw_col is not None and row[type1_raw_col] else ''
        type1_code = str(row[code_col]).strip() if code_col is not None and row[code_col] else ''
        type2 = str(row[type2_col]).strip() if type2_col is not None and row[type2_col] else ''
        type3 = str(row[type3_col]).strip() if type3_col is not None and row[type3_col] else ''
        special_type = str(row[special_col]).strip() if special_col is not None and row[special_col] else ''
        table[key] = {
            'type1_raw': type1_raw, 'type1': type1_code if type1_code else type1_raw,
            'type2': type2, 'type3': type3, 'special_type': special_type, 'raw_name': raw_name,
        }
    return table


# ─── 促销文本解析 ───

def parse_promo_text(text: str) -> list:
    if not text or str(text).strip() in ('', '无促销', 'nan', 'None'):
        return []
    results = []
    for line in str(text).split('\n'):
        line = line.strip()
        if not line:
            continue
        m = _PROMO_RE.match(line)
        if m:
            raw_name = m.group(1).strip()
            try:
                amount = float(m.group(2).replace(',', ''))
            except ValueError:
                amount = 0.0
        else:
            m2 = _PROMO_RE_FAN.match(line)
            if m2:
                raw_name = line  # 保留完整原始名称（含"返￥XX"）
                try:
                    amount = float(m2.group(2).replace(',', ''))
                except ValueError:
                    amount = 0.0
            else:
                raw_name = line
                amount = 0.0
        results.append({'raw_name': raw_name, 'name_norm': normalize_promo_name(raw_name), 'amount': amount})
    return results


def _trailing_digit_fallback(name_norm: str, code_table: dict):
    """尾部数字剥离兜底：积分当钱花23 → 积分当钱花，返回 (stripped_key, digit_amount)"""
    m = re.search(r'^(.*\D)(\d+(?:\.\d+)?)$', name_norm)
    if m:
        stripped = m.group(1).strip()
        if stripped and stripped in code_table:
            try:
                return stripped, float(m.group(2))
            except ValueError:
                pass
    return None, None


def enrich_q_promos(promos: list, code_table: dict) -> list:
    for p in promos:
        info = code_table.get(p['name_norm'], {})
        if not info:
            fallback_key, digit_amt = _trailing_digit_fallback(p['name_norm'], code_table)
            if fallback_key:
                info = code_table[fallback_key]
                if p['amount'] == 0.0 and digit_amt:
                    p['amount'] = digit_amt
        p['type1'] = info.get('type1', '')
        p['type2'] = info.get('type2', '')
        p['type3'] = info.get('type3', '')
        p['special_type'] = info.get('special_type', '')
        p['is_billion'] = info.get('is_billion', False)
        p['matched'] = p['name_norm'] in code_table or bool(info)
    return promos


def enrich_c_promos(promos: list, code_table: dict, promo_code_set: str = "xiecheng") -> list:
    for p in promos:
        info = code_table.get(p['name_norm'], {})
        if not info:
            fallback_key, digit_amt = _trailing_digit_fallback(p['name_norm'], code_table)
            if fallback_key:
                info = code_table[fallback_key]
                if p['amount'] == 0.0 and digit_amt:
                    p['amount'] = digit_amt
        type1_raw = info.get('type1_raw', '')
        type1_code = info.get('type1', '')
        special_type = info.get('special_type', '')
        if promo_code_set == "tongcheng":
            if p['raw_name'] == '可用券':
                special_type = '大额券' if p.get('amount', 0) >= 50 else '常规平台券'
            elif p['raw_name'] == '黑鲸优惠':
                special_type = '黑鲸优惠'
        p['type1_raw'] = type1_raw
        p['type1'] = type1_code if type1_code else type1_raw
        p['type2'] = info.get('type2', '')
        p['type3'] = info.get('type3', '')
        p['special_type'] = special_type
        p['matched'] = (p['name_norm'] in code_table or bool(info)) and p['type1'] not in ('', '未知')
    return promos


# ─── 校验逻辑 ───

def _safe_float(v):
    try:
        return float(v) if v not in (None, '', 'nan', 'None') else None
    except (ValueError, TypeError):
        return None


def validate_v1_amount(row: dict, q_promos: list, c_promos: list) -> list:
    errors = []
    q_line = _safe_float(row.get(FIELDS["q_list_price"]))
    q_pay = _safe_float(row.get(FIELDS["q_pay_price"]))
    c_line = _safe_float(row.get(FIELDS["c_list_price"]))
    c_pay = _safe_float(row.get(FIELDS["c_pay_price"]))

    if q_line is not None and q_pay is not None:
        q_total = sum(p['amount'] for p in q_promos)
        diff = abs(q_line - (q_total + q_pay))
        if diff > 1.0:
            errors.append(f'V1错误：Q金额不平衡（差值={diff:.2f}）')

    if c_line is not None and c_pay is not None:
        c_total = sum(p['amount'] for p in c_promos)
        diff = abs(c_line - (c_total + c_pay))
        if diff > 1.0:
            errors.append(f'V1错误：竞品金额不平衡（差值={diff:.2f}）')

    return errors


def validate_v2_promo_match(q_promos: list, c_promos: list) -> list:
    errors = []
    q_unmatched = [p['raw_name'] for p in q_promos if not p['matched']]
    c_unmatched = [p['raw_name'] for p in c_promos if not p['matched']]

    if q_unmatched:
        errors.append(f'V2错误：Q促销未匹配-{"，".join(q_unmatched)}')
    if c_unmatched:
        errors.append(f'V2错误：竞品促销未匹配-{"，".join(c_unmatched)}')
    return errors


def check_screenshot_platform(row: dict) -> bool:
    """返回True表示比价平台可以比价，False表示不可比价"""
    platform = str(row.get(FIELDS["compare_platform"], '') or '').strip()
    if not platform or platform in ('', 'nan', 'None'):
        return True
    return platform in VALID_PLATFORMS


def is_noncomparable_by_kf(kf_text: str) -> bool:
    """根据客服判断场景字段判断是否为不可对比场景。命中则跳过V1/V2校验和促销分行写入。"""
    if not kf_text:
        return False
    kf_lower = kf_text.lower()
    if 'roomid' in kf_lower:
        return True
    return any(kw in kf_text for kw in _KF_NONCOMPARABLE)


# ─── 促销数据文件 ───

def make_promo_output_path(input_path: str) -> str:
    dir_name = os.path.dirname(os.path.abspath(input_path))
    base = os.path.splitext(os.path.basename(input_path))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    return os.path.join(dir_name, f"{base}_促销数据_{timestamp}.xlsx")


def create_promo_file(promo_path: str, q_promo_rows: list, c_promo_rows: list):
    """创建独立的促销数据文件，含2个sheet"""
    openpyxl = check_openpyxl()
    wb = openpyxl.Workbook()

    ws_q = wb.active
    ws_q.title = "Q促销分行表"
    q_headers = ['平台', 'promo_code_set', 'case编号', '工单号', '促销名称', '促销金额',
                 '促销类型1级', '促销类型2级', '促销类型3级', '特殊类型', '是否十亿补贴',
                 'V1校验', 'V2校验']
    ws_q.append(q_headers)
    for row in q_promo_rows:
        ws_q.append(row)

    ws_c = wb.create_sheet("竞品促销分行表")
    c_headers = ['平台', 'promo_code_set', 'case编号', '工单号', '促销名称', '促销金额',
                 '促销类型1级', '促销类型2级', '促销类型3级', '特殊类型', 'V1校验', 'V2校验']
    ws_c.append(c_headers)
    for row in c_promo_rows:
        ws_c.append(row)

    wb.save(promo_path)


# ─── 主流程 ───

def _log(msg: str):
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", file=sys.stderr)


def main():
    _log("validate.py 启动")
    parser = argparse.ArgumentParser(description='促销校验模块')
    parser.add_argument('--input-file', required=False, default=None, help='原始输入Excel文件路径（含价格和促销数据）')
    parser.add_argument('--promo-file', required=False, default=None, help='促销数据输出文件路径（不填则自动生成）')
    parser.add_argument('--state-file', required=False, default=None, help='Orchestrator state.json 路径')
    parser.add_argument('--platform', required=False, default=None, help='本次运行平台：携程/同程')
    parser.add_argument('--promo-code-set', required=False, default=None, help='促销码表集合：xiecheng/tongcheng')
    parser.add_argument('--code-q', required=False, default=None, help='Q促销码表路径（兼容旧调用）')
    parser.add_argument('--code-c', required=False, default=None, help='竞品促销码表路径（兼容旧调用）')
    args = parser.parse_args()

    state_input = load_state_options(args.state_file)
    platform = require_platform(args.platform or state_input.get('platform') or state_input.get('归因平台'))
    promo_code_set = normalize_rule_set(args.promo_code_set or state_input.get('promo_code_set'), platform)
    code_q_path, code_c_path, q_sheet_name, c_sheet_name = resolve_code_table_paths(args, platform, promo_code_set)

    # 加载字段名映射（field_map.json）
    loaded_fields = load_field_map(promo_code_set)
    if loaded_fields:
        FIELDS.update(loaded_fields)
        _log(f"字段映射已加载（promo_code_set={promo_code_set}，共 {len(loaded_fields)} 项）")
    else:
        _log("警告：field_map.json 未找到，使用默认字段名")

    # 确定原始 Excel 路径
    input_file = args.input_file or state_input.get('excel')
    if not input_file:
        print("错误：缺少原始Excel路径。请通过 --input-file 或 state.json 的 input.excel 字段提供", file=sys.stderr)
        sys.exit(1)
    if not os.path.exists(input_file):
        print(f"错误：文件不存在 - {input_file}", file=sys.stderr)
        sys.exit(1)

    # 确定促销数据输出路径
    promo_output_path = args.promo_file or make_promo_output_path(input_file)

    for p in [code_q_path, code_c_path]:
        if not os.path.exists(p):
            print(f"错误：文件不存在 - {p}", file=sys.stderr)
            sys.exit(1)

    openpyxl = check_openpyxl()

    _log(f"运行平台: {platform} / promo_code_set={promo_code_set}")
    _log(f"加载Q码表: {code_q_path}" + (f"#{q_sheet_name}" if q_sheet_name else ""))
    code_q = load_code_table_q(code_q_path, q_sheet_name)
    _log(f"  → Q码表 {len(code_q)} 条")

    _log(f"加载竞品码表: {code_c_path}" + (f"#{c_sheet_name}" if c_sheet_name else ""))
    code_c = load_code_table_c(code_c_path, c_sheet_name)
    _log(f"  → 竞品码表 {len(code_c)} 条")

    _log(f"读取原始输入文件: {input_file}")
    wb_input = openpyxl.load_workbook(input_file)
    if "主数据" in wb_input.sheetnames:
        ws_main = wb_input["主数据"]
    else:
        ws_main = wb_input.worksheets[0]

    headers = [ws_main.cell(1, c).value for c in range(1, ws_main.max_column + 1)]
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i + 1

    results = []
    q_promo_rows = []
    c_promo_rows = []

    for row_num in range(2, ws_main.max_row + 1):
        def get_cell(col_name, _row_num=row_num):
            col = col_map.get(col_name)
            if col is None:
                return None
            v = ws_main.cell(_row_num, col).value
            return str(v).strip() if v is not None else None

        case_id = get_cell(FIELDS["case_no"]) or f"row_{row_num}"
        flow_no = get_cell(FIELDS["order_no"]) or ''
        kf_text = get_cell(FIELDS["kf_note"]) or ''
        row_dict = {h: get_cell(str(h)) for h in headers if h}

        # 不可对比场景判断（基于客服判断场景，优先级最高）
        if is_noncomparable_by_kf(kf_text):
            results.append({
                'row_num': row_num,
                'case_id': case_id,
                'screenshot_platform_ok': True,
                'v1_pass': True,
                'v2_pass': True,
                'errors': [],
                'skipped': True,
                'skip_reason': '不可对比场景（客服判断场景判断），跳过校验',
            })
            continue

        q_promo_text = row_dict.get(FIELDS["q_promo_detail"]) or ''
        c_promo_text = row_dict.get(FIELDS["c_promo_detail"]) or ''

        q_promos = enrich_q_promos(parse_promo_text(q_promo_text), code_q)
        c_promos = enrich_c_promos(parse_promo_text(c_promo_text), code_c, promo_code_set)

        platform_ok = check_screenshot_platform(row_dict)
        v1_errors = validate_v1_amount(row_dict, q_promos, c_promos)
        v2_errors = validate_v2_promo_match(q_promos, c_promos)

        all_errors = v1_errors + v2_errors
        if not platform_ok:
            actual_platform = row_dict.get(FIELDS["compare_platform"]) or ''
            all_errors = [f'比价平台不可比价：{actual_platform}'] + all_errors

        results.append({
            'row_num': row_num,
            'case_id': case_id,
            'screenshot_platform_ok': platform_ok,
            'v1_pass': len(v1_errors) == 0,
            'v2_pass': len(v2_errors) == 0,
            'errors': all_errors,
            'skipped': False,
        })

        q_v1_note = '；'.join(e for e in v1_errors if 'Q' in e)
        c_v1_note = '；'.join(e for e in v1_errors if '竞品金额' in e)

        for p in q_promos:
            q_promo_rows.append([
                platform, promo_code_set, case_id, flow_no,
                p['raw_name'], p['amount'],
                p.get('type1', ''), p.get('type2', ''), p.get('type3', ''),
                p.get('special_type', ''),
                '是' if p.get('is_billion') else '否',
                q_v1_note,
                '' if p.get('matched') else '未匹配',
            ])

        for p in c_promos:
            c_promo_rows.append([
                platform, promo_code_set, case_id, flow_no,
                p['raw_name'], p['amount'],
                p.get('type1_raw', '') or p.get('type1', ''),
                p.get('type2', ''), p.get('type3', ''),
                p.get('special_type', ''),
                c_v1_note,
                '' if p.get('matched') else '未匹配',
            ])

    # 创建促销数据文件
    _log(f"创建促销数据文件: {promo_output_path}")
    create_promo_file(promo_output_path, q_promo_rows, c_promo_rows)
    _log(f"Q促销分行表 写入 {len(q_promo_rows)} 条，竞品促销分行表 写入 {len(c_promo_rows)} 条")

    # 构建 case_audit（每 case 的人工校验原因，供归因模块写入主数据）
    case_audit = {}
    for r in results:
        reasons = []
        for err in r.get('errors', []):
            if err.startswith('V1错误') or err.startswith('V2错误'):
                reasons.append(f'真实错误:{err}')
        if reasons:
            case_audit[r['case_id']] = reasons

    # 更新 state.json
    if args.state_file and os.path.exists(args.state_file):
        with open(args.state_file, 'r', encoding='utf-8') as f:
            state = json.load(f)

        skipped_items = [r for r in results if r.get('skipped')]
        checked_items = [r for r in results if not r.get('skipped')]
        v1_fails = [r for r in checked_items if not r['v1_pass']]
        v2_fails = [r for r in checked_items if not r['v2_pass']]
        platform_fails = [r for r in checked_items if not r['screenshot_platform_ok']]

        state["agents"]["promo"]["status"] = "done"
        state["agents"]["promo"]["completed_at"] = datetime.now().isoformat()
        state["agents"]["promo"]["promo_file"] = promo_output_path
        state["agents"]["promo"]["validation_summary"] = {
            "total_rows": len(results),
            "v1_pass": len(checked_items) - len(v1_fails),
            "v1_fail": len(v1_fails),
            "v2_pass": len(checked_items) - len(v2_fails),
            "v2_fail": len(v2_fails),
            "platform_blocked": len(platform_fails),
            "manual_intervention_count": len(case_audit),
        }
        state["agents"]["promo"]["case_audit"] = case_audit
        state["promo_file"] = promo_output_path

        # 将 V1/V2 失败 case 写入 fallback pending_cases，供 Orchestrator 调度 LLM 兜底
        if case_audit:
            if "fallback" not in state["agents"]:
                state["agents"]["fallback"] = {"pending_cases": []}
            if "pending_cases" not in state["agents"]["fallback"]:
                state["agents"]["fallback"]["pending_cases"] = []
            pending = state["agents"]["fallback"]["pending_cases"]
            existing_ids = {c["case_id"] for c in pending}
            for cid, errors in case_audit.items():
                if cid not in existing_ids:
                    pending.append({
                        "case_id": cid,
                        "scenario": "promo_v2_unmatched",
                        "source": "promo",
                        "errors": errors,
                        "resolved": False,
                        "tier": None,
                    })
            state["agents"]["fallback"]["pending_cases"] = pending

        with open(args.state_file, 'w', encoding='utf-8') as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
        _log("state.json 已更新（E: done）")

    # 统计
    skipped_items = [r for r in results if r.get('skipped')]
    checked_items = [r for r in results if not r.get('skipped')]
    v1_fails = [r for r in checked_items if not r['v1_pass']]
    v2_fails = [r for r in checked_items if not r['v2_pass']]
    platform_fails = [r for r in checked_items if not r['screenshot_platform_ok']]
    print(f"不可对比场景（跳过校验）：{len(skipped_items)} 条", file=sys.stderr)
    print(f"参与校验：{len(checked_items)} 条", file=sys.stderr)
    print(f"V1校验：{len(checked_items)-len(v1_fails)}通过 / {len(v1_fails)}失败", file=sys.stderr)
    print(f"V2校验：{len(checked_items)-len(v2_fails)}通过 / {len(v2_fails)}失败", file=sys.stderr)
    print(f"比价平台不可比价：{len(platform_fails)} 条", file=sys.stderr)
    print(f"人工校验标记：{len(case_audit)} 条（写入 state.json case_audit）", file=sys.stderr)
    print(f"促销数据文件：{promo_output_path}", file=sys.stderr)

    _log(f"校验完成: 总数={len(results)}, 跳过={len(skipped_items)}, V1失败={len(v1_fails)}, V2失败={len(v2_fails)}")
    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
