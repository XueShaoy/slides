#!/usr/bin/env python3
"""
validate_fallback.py — 促销校验兜底处理

在 validate.py 完成基础解析后，对仍有 V1/V2 失败的 case 进行两级兜底：
  兜底2：读取【明细页-价格优惠明细页】URL，Claude Vision 识别竞品截图内容（优先）
  兜底1：LLM 重新解析原始促销文本（兜底2失败后执行）

执行顺序：兜底2先（截图可绕过OCR错误和分行格式问题）→ 兜底1后（文本推断）
Q侧和竞品侧独立处理，各自的建议只写各自的分行表。

V1（金额平衡）和 V2（促销类型匹配）独立判断：
  - V2：兜底重新解析名称后，跑码表 enrich 匹配，all(matched) = V2解决
  - V1：用新金额做平衡校验，通过 = V1解决

兜底方式列四种状态（每侧独立）：
  兜底x              → V1 ✅ V2 ✅
  兜底x-V1待处理      → V1 ❌ V2 ✅（金额仍不平衡，类型已识别）
  兜底x-V2待码表更新   → V1 ✅ V2 ❌（金额平衡，但名称仍未入码表）
  需人工介入          → V1 ❌ V2 ❌

Usage:
  python3 validate_fallback.py \
    --state-file <state.json路径> \
    --input-file <原始Excel路径>
"""

import argparse
import base64
import difflib
import json
import os
import re
import sys
from datetime import datetime

import anthropic
import openpyxl
import requests

# 从 validate.py 导入码表加载和促销解析函数
sys.path.insert(0, os.path.dirname(__file__))
from validate import (
    normalize_promo_name,
    enrich_q_promos,
    enrich_c_promos,
    load_code_table_q,
    load_code_table_c,
    parse_promo_text,
)

# ─── 常量 ───────────────────────────────────────────────────────────────────

CLAUDE_MODEL = "claude-sonnet-4-6"
IMAGE_DETAIL_FIELD = "明细页-价格优惠明细页"
Q_PROMO_FIELD = "Q优惠明细"
C_PROMO_FIELD = "促销明细（竞品）"
CASE_NO_FIELD = "case编号"

CODE_TABLE_PATH_TONGCHENG = "/Users/zhangwang/.claude/agents/orion-bijiaups-rca-promo/code_tables/E_QE.xlsx"
CODE_TABLE_PATH_Q_DEFAULT  = "/Users/zhangwang/.claude/agents/orion-bijiaups-rca-promo/code_tables/C_Q.xlsx"
CODE_TABLE_PATH_C_DEFAULT  = "/Users/zhangwang/.claude/agents/orion-bijiaups-rca-promo/code_tables/C_C.xlsx"

# ─── 工具函数 ─────────────────────────────────────────────────────────────────

def _log(msg: str):
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}", file=sys.stderr)


def _load_state(state_file: str) -> dict:
    with open(state_file, encoding="utf-8") as f:
        return json.load(f)


def _save_state(state_file: str, state: dict):
    with open(state_file, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def _fetch_image_base64(url: str) -> tuple[str, str]:
    resp = requests.get(url, timeout=15)
    resp.raise_for_status()
    ct = resp.headers.get("Content-Type", "image/jpeg").split(";")[0].strip()
    if ct not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
        ct = "image/jpeg"
    return base64.standard_b64encode(resp.content).decode("utf-8"), ct


def _call_claude(client: anthropic.Anthropic, messages: list, max_tokens: int = 1024) -> str:
    resp = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=max_tokens,
        messages=messages,
    )
    return resp.content[0].text


def _extract_json(text: str) -> dict | list | None:
    m = re.search(r"```json\s*([\s\S]+?)\s*```", text)
    if m:
        text = m.group(1)
    else:
        m = re.search(r"(\{[\s\S]+\}|\[[\s\S]+\])", text)
        if m:
            text = m.group(1)
    try:
        return json.loads(text)
    except Exception:
        return None


def _load_code_tables(promo_code_set: str) -> tuple[dict, dict]:
    """根据 promo_code_set 加载 Q码表 和 竞品码表"""
    if promo_code_set == "tongcheng":
        code_q = load_code_table_q(CODE_TABLE_PATH_TONGCHENG, sheet_name="Q促销码表")
        code_c = load_code_table_c(CODE_TABLE_PATH_TONGCHENG, sheet_name="E促销码表")
    else:
        code_q = load_code_table_q(CODE_TABLE_PATH_Q_DEFAULT)
        code_c = load_code_table_c(CODE_TABLE_PATH_C_DEFAULT)
    return code_q, code_c


# ─── 兜底2：Claude Vision 识别竞品截图（优先）────────────────────────────────

FALLBACK2_PROMPT = """请识别这张酒店订单截图中的「价格优惠明细」或「促销明细」部分。
提取每条优惠的名称和金额（元），如果截图没有明细信息请返回空列表。

返回 JSON（无其他文字）：
{
  "promos": [
    {"name": "优惠名称", "amount": 金额数字},
    ...
  ],
  "optimization": "截图来源平台及促销特征，建议如何完善竞品码表（一句话）"
}
"""

def fallback2_vision(client: anthropic.Anthropic, image_url: str) -> dict:
    """兜底2：Vision 识别竞品截图（仅处理竞品侧）"""
    try:
        img_b64, media_type = _fetch_image_base64(image_url)
        reply = _call_claude(client, [{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
                {"type": "text", "text": FALLBACK2_PROMPT},
            ],
        }], max_tokens=1024)
        parsed = _extract_json(reply)
        if parsed and isinstance(parsed, dict):
            return parsed
    except Exception as e:
        _log(f"  兜底2 Vision 调用失败: {e}")
    return {"promos": [], "optimization": ""}


# ─── 兜底1：LLM 解析原始促销文本（单侧）──────────────────────────────────────

FALLBACK1_PROMPT = """你是促销解析专家。请从以下促销明细文本中，提取每条促销的名称和金额（元，数字类型）。

规则：
1. 格式多样：「名称 -¥金额」「名称：金额」「名称 返￥金额」「名称金额（如"积分当钱花23"表示23元）」
2. 如果文本本身就是「#N/A」或空，返回空列表
3. 每条促销单独一项，不合并

请返回 JSON（无其他文字）：
{
  "promos": [
    {"name": "促销名称", "amount": 金额数字},
    ...
  ],
  "optimization": "发现了什么新的命名模式，建议如何扩展解析规则或码表（一句话）"
}

促销文本：
%TEXT%
"""

def fallback1_llm_single(client: anthropic.Anthropic, text: str) -> dict:
    """兜底1：解析单侧促销文本"""
    if not text or str(text).strip() in ("", "nan", "None", "#N/A"):
        return {"promos": [], "optimization": ""}
    prompt = FALLBACK1_PROMPT.replace("%TEXT%", str(text).strip())
    try:
        reply = _call_claude(client, [{"role": "user", "content": prompt}])
        parsed = _extract_json(reply)
        if parsed and isinstance(parsed, dict):
            return {"promos": parsed.get("promos", []), "optimization": parsed.get("optimization", "")}
    except Exception as e:
        _log(f"  兜底1 LLM 调用失败: {e}")
    return {"promos": [], "optimization": ""}


# ─── 促销解析结果转标准格式 ───────────────────────────────────────────────────

def _to_promo_dicts(raw_promos: list) -> list:
    """把兜底返回的 [{"name":..., "amount":...}] 转为 validate.py enrich 所需格式"""
    result = []
    for p in raw_promos:
        name = str(p.get("name", "")).strip()
        if not name:
            continue
        result.append({
            "raw_name": name,
            "name_norm": normalize_promo_name(name),
            "amount": float(p["amount"]) if isinstance(p.get("amount"), (int, float)) else 0.0,
        })
    return result


# ─── V1 重校验（各侧独立）────────────────────────────────────────────────────

def _get_num(row_dict: dict, field: str) -> float:
    v = row_dict.get(field)
    try:
        return float(str(v).replace(",", "")) if v not in (None, "", "nan", "None") else 0.0
    except Exception:
        return 0.0


def recheck_v1_q(row_dict: dict, promos: list) -> bool:
    """Q侧 V1 重校验，返回 True = 通过"""
    q_listed = _get_num(row_dict, "Q折后底价——同质化房型") or _get_num(row_dict, "Q折后底价——物理房型最低价")
    q_paid   = _get_num(row_dict, "Q到手价——同质化房型") or _get_num(row_dict, "Q到手价——物理房型最低价")
    if q_listed <= 0:
        return True  # 无划线价，跳过
    total = sum(p["amount"] for p in promos)
    return abs(q_listed - q_paid - total) <= 1.0


def recheck_v1_c(row_dict: dict, promos: list) -> bool:
    """竞品侧 V1 重校验，返回 True = 通过"""
    c_listed = _get_num(row_dict, "划线价（竞品）")
    c_paid   = _get_num(row_dict, "到手价（竞品）")
    if c_listed <= 0:
        return True  # 无划线价，跳过
    total = sum(p["amount"] for p in promos)
    return abs(c_listed - c_paid - total) <= 1.0


def check_v2(promos: list) -> bool:
    """V2 校验：所有促销均已匹配码表，返回 True = 通过"""
    if not promos:
        return False
    return all(p.get("matched", False) for p in promos)


FUZZY_MATCH_CUTOFF = 0.75


def _describe_unmatched_name(raw_name: str, code_table: dict) -> str:
    """
    对码表未命中的促销名称做归因说明：
      - 名称本身已在码表中（只是类型未分类）→ 提示补全类型，不是OCR问题
      - 名称与码表某条目高度相似（编辑距离小）→ 疑似OCR/形近字识别错误，建议核对原始截图
      - 均不满足 → 视为码表确实未收录该促销
    """
    name_norm = normalize_promo_name(raw_name)
    info = code_table.get(name_norm)
    if info:
        return f"「{raw_name}」已在码表中，但促销类型未分类，建议人工补全码表类型"

    candidates = list(code_table.keys())
    matches = difflib.get_close_matches(name_norm, candidates, n=1, cutoff=FUZZY_MATCH_CUTOFF)
    if matches:
        best = matches[0]
        ratio = difflib.SequenceMatcher(None, name_norm, best).ratio()
        return f"「{raw_name}」疑似OCR/形近字识别错误，与码表「{best}」相似度{ratio:.0%}，建议核对原始截图后更正名称"

    return f"「{raw_name}」未命中码表，建议确认是否为新促销并补充码表"


# ─── 兜底结果：单侧处理 ───────────────────────────────────────────────────────

def _resolve_side(promos: list, row_dict: dict, side: str,
                  code_q: dict, code_c: dict, promo_code_set: str,
                  fb_source: str) -> dict:
    """
    对单侧（q/c）兜底结果做 V1+V2 独立判断。
    返回 {
        "tier": str,          # 兜底方式列写入值
        "v1_resolved": bool,
        "v2_resolved": bool,
        "enriched": list,     # enrich 后的 promo 列表（供写回分行表）
        "amounts_resolved": list,  # 用于更新促销金额的列表（v1解决时才有）
    }
    """
    if not promos:
        return {"tier": "需人工介入", "v1_resolved": False, "v2_resolved": False,
                "enriched": [], "amounts_resolved": None}

    # 转标准格式
    promo_dicts = _to_promo_dicts(promos)

    # 码表 enrich → 判断 V2
    if side == "q":
        enriched = enrich_q_promos(promo_dicts, code_q)
    else:
        enriched = enrich_c_promos(promo_dicts, code_c, promo_code_set)

    v2_ok = check_v2(enriched)
    v1_ok = recheck_v1_q(row_dict, enriched) if side == "q" else recheck_v1_c(row_dict, enriched)

    if v1_ok and v2_ok:
        tier = fb_source
    elif not v1_ok and v2_ok:
        tier = f"{fb_source}-V1待处理"
    elif v1_ok and not v2_ok:
        tier = f"{fb_source}-V2待码表更新"
    else:
        # 兜底解析出来了但 V1/V2 都没解决，仍标"需人工介入"
        tier = "需人工介入"

    amounts_resolved = enriched if v1_ok else None

    return {
        "tier": tier,
        "v1_resolved": v1_ok,
        "v2_resolved": v2_ok,
        "enriched": enriched,
        "amounts_resolved": amounts_resolved,
    }


# ─── 更新促销分行表 ────────────────────────────────────────────────────────────

def update_promo_file(promo_file: str, case_id: str, side: str,
                      tier: str, optimization: str,
                      v1_resolved: bool, v2_resolved: bool,
                      amounts_resolved: list | None = None):
    """
    在促销分行表对应 case 的行上写入兜底结果。

    V1/V2 独立处理：
      V1 解决 → V1校验列写"通过兜底x已解决"（x = 兜底1/2，不含后缀）
      V2 解决 → V2校验列"未匹配"行改为"通过兜底x已解决"
      未解决  → 对应列保持原样不动

    optimization 只写入该侧分行表，不跨侧写。
    """
    wb = openpyxl.load_workbook(promo_file)
    sheet_name = "Q促销分行表" if side == "q" else "竞品促销分行表"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return

    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def _ensure_col(name):
        if name not in headers:
            idx = ws.max_column + 1
            ws.cell(1, idx).value = name
            headers.append(name)
        return headers.index(name) + 1

    tier_col   = _ensure_col("兜底方式")
    opt_col    = _ensure_col("兜底优化建议")
    case_col   = headers.index("case编号") + 1 if "case编号" in headers else None
    if case_col is None:
        wb.save(promo_file)
        return

    amount_col = headers.index("促销金额") + 1 if "促销金额" in headers else None
    v1_col     = headers.index("V1校验") + 1   if "V1校验" in headers else None
    v2_col     = headers.index("V2校验") + 1   if "V2校验" in headers else None

    # 兜底方式列写入的"基础名称"（去掉后缀，用于标注V1/V2列）
    base_fb = tier.split("-")[0]  # "兜底1" 或 "兜底2" 或 "需人工介入"

    case_rows = [row for row in ws.iter_rows(min_row=2)
                 if str(row[case_col - 1].value or "").strip() == str(case_id).strip()]

    for idx, row in enumerate(case_rows):
        row_num = row[0].row

        # 兜底方式（所有情况都写）
        ws.cell(row_num, tier_col).value = tier

        # 优化建议（有内容才写）
        if optimization:
            ws.cell(row_num, opt_col).value = optimization

        # V1：解决了才改，保持原样
        if v1_resolved and v1_col:
            orig_v1 = ws.cell(row_num, v1_col).value
            if orig_v1:
                ws.cell(row_num, v1_col).value = f"通过{base_fb}已解决"

        # V2：解决了才改，未解决保持"未匹配"
        if v2_resolved and v2_col:
            orig_v2 = ws.cell(row_num, v2_col).value
            if orig_v2 == "未匹配":
                ws.cell(row_num, v2_col).value = f"通过{base_fb}已解决"

        # 促销金额：V1 解决时按行覆写
        if amounts_resolved and amount_col and idx < len(amounts_resolved):
            new_amount = amounts_resolved[idx].get("amount")
            if new_amount is not None:
                ws.cell(row_num, amount_col).value = new_amount

    wb.save(promo_file)


# ─── 批量脚本兜底（默认入口，不依赖 ANTHROPIC_API_KEY）────────────────────────

INVALID_PROMO_NAMES = {"", "#N/A", "0", "nan", "None", "无", "无促销"}


def _is_invalid_promo_name(name: object) -> bool:
    return str(name or "").strip() in INVALID_PROMO_NAMES


def _to_float(value: object) -> float:
    try:
        if value in (None, "", "nan", "None"):
            return 0.0
        return float(str(value).replace(",", "").strip())
    except Exception:
        return 0.0


def _side_has_error(errors: list, side: str) -> bool:
    if side == "q":
        return any("Q促销" in e or "Q金额" in e for e in errors)
    return any("竞品促销" in e or "竞品金额" in e for e in errors)


def _read_input_rows(input_file: str, target_case_ids: set[str]) -> dict:
    wb = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    ws = wb["主数据"] if "主数据" in wb.sheetnames else wb.worksheets[0]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {str(h).strip(): i for i, h in enumerate(headers) if h}
    case_col = col_map.get(CASE_NO_FIELD)

    rows = {}
    for row_idx, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_case = str(cells[case_col] or "").strip() if case_col is not None else ""
        case_id = raw_case if raw_case else f"row_{row_idx}"
        if case_id not in target_case_ids:
            continue
        rows[case_id] = {
            str(h).strip(): (cells[i] if i < len(cells) else None)
            for i, h in enumerate(headers) if h
        }
    wb.close()
    return rows


def _load_side_rows(promo_file: str, side: str) -> dict:
    wb = openpyxl.load_workbook(promo_file, read_only=True, data_only=True)
    sheet_name = "Q促销分行表" if side == "q" else "竞品促销分行表"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb[sheet_name]
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows_by_case = {}
    for cells in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: cells[i] if i < len(cells) else None for i in range(len(headers))}
        case_id = str(item.get("case编号") or "").strip()
        if not case_id:
            continue
        rows_by_case.setdefault(case_id, []).append(item)
    wb.close()
    return rows_by_case


def _promos_from_sheet_rows(rows: list[dict]) -> list:
    promos = []
    for row in rows:
        raw_name = str(row.get("促销名称") or "").strip()
        amount = _to_float(row.get("促销金额"))
        if _is_invalid_promo_name(raw_name):
            continue

        # 兼容「普卡特权;87」这类名称金额粘在一起的格式。
        m = re.match(r"^(.+?)[;；]\s*([+-]?\d+(?:\.\d+)?)$", raw_name)
        if m:
            raw_name = m.group(1).strip()
            if amount == 0.0:
                amount = float(m.group(2))

        promos.append({
            "raw_name": raw_name,
            "name_norm": normalize_promo_name(raw_name),
            "amount": amount,
        })
    return promos


def _resolve_sheet_side(rows: list[dict], row_dict: dict, side: str,
                        code_q: dict, code_c: dict, promo_code_set: str) -> dict:
    promos = _promos_from_sheet_rows(rows)
    if not promos:
        return {"tier": "需人工介入", "v1_resolved": False, "v2_resolved": False,
                "enriched": [], "amounts_resolved": None,
                "optimization": "原始促销明细为空、#N/A或0，无可解析文本"}

    if side == "q":
        enriched = enrich_q_promos(promos, code_q)
        v1_ok = recheck_v1_q(row_dict, enriched)
    else:
        enriched = enrich_c_promos(promos, code_c, promo_code_set)
        v1_ok = recheck_v1_c(row_dict, enriched)

    v2_ok = check_v2(enriched)

    if v1_ok and v2_ok:
        tier = "兜底1"
    elif not v1_ok and v2_ok:
        tier = "兜底1-V1待处理"
    elif v1_ok and not v2_ok:
        tier = "兜底1-V2待码表更新"
    else:
        tier = "需人工介入"

    unmatched_messages = {}
    if tier.endswith("V2待码表更新"):
        code_table = code_q if side == "q" else code_c
        for p in enriched:
            if not p.get("matched"):
                unmatched_messages[p["raw_name"]] = _describe_unmatched_name(p["raw_name"], code_table)
        optimization = "；".join(unmatched_messages.values())
    elif tier == "兜底1":
        optimization = "通过批量脚本兜底重新解析名称/金额并命中码表"
    else:
        optimization = ""

    return {
        "tier": tier,
        "v1_resolved": v1_ok,
        "v2_resolved": v2_ok,
        "enriched": enriched,
        "amounts_resolved": enriched if v1_ok else None,
        "optimization": optimization,
        "unmatched_messages": unmatched_messages,
    }


def _write_side_batch(promo_file: str, case_id: str, side: str, result: dict):
    wb = openpyxl.load_workbook(promo_file)
    sheet_name = "Q促销分行表" if side == "q" else "竞品促销分行表"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return

    ws = wb[sheet_name]
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def ensure_col(name):
        if name not in headers:
            idx = ws.max_column + 1
            ws.cell(1, idx).value = name
            headers.append(name)
        return headers.index(name) + 1

    col = {h: i + 1 for i, h in enumerate(headers)}
    tier_col = ensure_col("兜底方式")
    opt_col = ensure_col("兜底优化建议")
    col = {h: i + 1 for i, h in enumerate(headers)}
    case_col = col.get("case编号")
    if not case_col:
        wb.save(promo_file)
        return

    case_rows = [row for row in ws.iter_rows(min_row=2)
                 if str(row[case_col - 1].value or "").strip() == str(case_id).strip()]
    enriched = result.get("enriched") or []
    base_fb = result["tier"].split("-")[0]

    for idx, row in enumerate(case_rows):
        row_num = row[0].row
        ws.cell(row_num, tier_col).value = result["tier"]
        if result.get("optimization"):
            ws.cell(row_num, opt_col).value = result["optimization"]

        if idx < len(enriched):
            p = enriched[idx]
            if "促销名称" in col:
                ws.cell(row_num, col["促销名称"]).value = p.get("raw_name", "")
            if "促销金额" in col:
                ws.cell(row_num, col["促销金额"]).value = p.get("amount", 0.0)
            if p.get("matched"):
                if "促销类型1级" in col:
                    ws.cell(row_num, col["促销类型1级"]).value = p.get("type1_raw") or p.get("type1", "")
                if "促销类型2级" in col:
                    ws.cell(row_num, col["促销类型2级"]).value = p.get("type2", "")
                if "促销类型3级" in col:
                    ws.cell(row_num, col["促销类型3级"]).value = p.get("type3", "")
                if "特殊类型" in col:
                    ws.cell(row_num, col["特殊类型"]).value = p.get("special_type", "")
                if "是否十亿补贴" in col:
                    ws.cell(row_num, col["是否十亿补贴"]).value = "是" if p.get("is_billion") else "否"

        if result.get("v1_resolved") and "V1校验" in col:
            orig_v1 = ws.cell(row_num, col["V1校验"]).value
            if orig_v1:
                ws.cell(row_num, col["V1校验"]).value = f"通过{base_fb}已解决"

        if result.get("v2_resolved") and "V2校验" in col:
            orig_v2 = ws.cell(row_num, col["V2校验"]).value
            if orig_v2 == "未匹配":
                ws.cell(row_num, col["V2校验"]).value = f"通过{base_fb}已解决"

    wb.save(promo_file)


def _prepare_write_context(wb, side: str) -> dict | None:
    sheet_name = "Q促销分行表" if side == "q" else "竞品促销分行表"
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def ensure_col(name):
        if name not in headers:
            idx = ws.max_column + 1
            ws.cell(1, idx).value = name
            headers.append(name)
        return headers.index(name) + 1

    ensure_col("兜底方式")
    ensure_col("兜底优化建议")
    col = {h: i + 1 for i, h in enumerate(headers)}
    case_col = col.get("case编号")
    if not case_col:
        return None

    rows_by_case = {}
    for row in ws.iter_rows(min_row=2):
        case_id = str(row[case_col - 1].value or "").strip()
        if case_id:
            rows_by_case.setdefault(case_id, []).append(row)

    return {"ws": ws, "col": col, "rows_by_case": rows_by_case}


def _write_side_batch_context(ctx: dict | None, case_id: str, result: dict):
    if not ctx:
        return

    ws = ctx["ws"]
    col = ctx["col"]
    case_rows = ctx["rows_by_case"].get(str(case_id).strip(), [])
    enriched = result.get("enriched") or []
    base_fb = result["tier"].split("-")[0]
    unmatched_messages = result.get("unmatched_messages") or {}
    # 纯 V2 未命中码表时，兜底方式/优化建议只标真正未匹配的那一行，
    # 避免同一 case 下已正常匹配的促销行也被打上同样的标注。
    only_unmatched_rows = result["tier"].endswith("V2待码表更新")

    for idx, row in enumerate(case_rows):
        row_num = row[0].row
        p = enriched[idx] if idx < len(enriched) else None
        is_unmatched_row = bool(p) and not p.get("matched")

        if only_unmatched_rows:
            if is_unmatched_row:
                ws.cell(row_num, col["兜底方式"]).value = result["tier"]
                msg = unmatched_messages.get(p["raw_name"], result.get("optimization", ""))
                if msg:
                    ws.cell(row_num, col["兜底优化建议"]).value = msg
        else:
            ws.cell(row_num, col["兜底方式"]).value = result["tier"]
            if result.get("optimization"):
                ws.cell(row_num, col["兜底优化建议"]).value = result["optimization"]

        if idx < len(enriched):
            p = enriched[idx]
            if "促销名称" in col:
                ws.cell(row_num, col["促销名称"]).value = p.get("raw_name", "")
            if "促销金额" in col:
                ws.cell(row_num, col["促销金额"]).value = p.get("amount", 0.0)
            if p.get("matched"):
                if "促销类型1级" in col:
                    ws.cell(row_num, col["促销类型1级"]).value = p.get("type1_raw") or p.get("type1", "")
                if "促销类型2级" in col:
                    ws.cell(row_num, col["促销类型2级"]).value = p.get("type2", "")
                if "促销类型3级" in col:
                    ws.cell(row_num, col["促销类型3级"]).value = p.get("type3", "")
                if "特殊类型" in col:
                    ws.cell(row_num, col["特殊类型"]).value = p.get("special_type", "")
                if "是否十亿补贴" in col:
                    ws.cell(row_num, col["是否十亿补贴"]).value = "是" if p.get("is_billion") else "否"

        if result.get("v1_resolved") and "V1校验" in col:
            orig_v1 = ws.cell(row_num, col["V1校验"]).value
            if orig_v1:
                ws.cell(row_num, col["V1校验"]).value = f"通过{base_fb}已解决"

        if result.get("v2_resolved") and "V2校验" in col:
            orig_v2 = ws.cell(row_num, col["V2校验"]).value
            if orig_v2 == "未匹配":
                ws.cell(row_num, col["V2校验"]).value = f"通过{base_fb}已解决"


def _remaining_audit_errors(errors: list, q_result: dict | None, c_result: dict | None) -> list:
    remaining = []
    for err in errors:
        if _side_has_error([err], "q") and q_result:
            if q_result["tier"] in ("兜底1", "兜底2", "兜底1-V2待码表更新", "兜底2-V2待码表更新"):
                continue
        if _side_has_error([err], "c") and c_result:
            if c_result["tier"] in ("兜底1", "兜底2", "兜底1-V2待码表更新", "兜底2-V2待码表更新"):
                continue
        remaining.append(err)
    return remaining


def _case_tier(q_result: dict | None, c_result: dict | None) -> str:
    parts = []
    if c_result:
        parts.append(f"竞品:{c_result['tier']}" if q_result else c_result["tier"])
    if q_result:
        parts.append(f"Q侧:{q_result['tier']}" if c_result else q_result["tier"])
    return "; ".join(parts) if parts else ""


def run_batch_fallback(state: dict, state_file: str, input_file: str, promo_file: str) -> dict:
    case_audit = state.get("agents", {}).get("promo", {}).get("case_audit", {}) or {}
    if not case_audit:
        return {"total": 0, "resolved": 0, "v2_code_table_update": 0, "needs_human": 0}

    promo_code_set = state.get("input", {}).get("promo_code_set", "xiecheng")
    code_q, code_c = _load_code_tables(promo_code_set)
    target_case_ids = set(case_audit.keys())
    input_rows = _read_input_rows(input_file, target_case_ids)
    q_rows = _load_side_rows(promo_file, "q")
    c_rows = _load_side_rows(promo_file, "c")
    wb_promo = openpyxl.load_workbook(promo_file)
    q_write_ctx = _prepare_write_context(wb_promo, "q")
    c_write_ctx = _prepare_write_context(wb_promo, "c")

    pending_cases = []
    new_case_audit = {}
    stats = {"total": len(case_audit), "resolved": 0, "v2_code_table_update": 0, "needs_human": 0}

    for case_id, errors in case_audit.items():
        row_dict = input_rows.get(case_id, {})
        q_result = None
        c_result = None

        if _side_has_error(errors, "q"):
            q_result = _resolve_sheet_side(q_rows.get(case_id, []), row_dict, "q", code_q, code_c, promo_code_set)
            _write_side_batch_context(q_write_ctx, case_id, q_result)

        if _side_has_error(errors, "c"):
            c_result = _resolve_sheet_side(c_rows.get(case_id, []), row_dict, "c", code_q, code_c, promo_code_set)
            _write_side_batch_context(c_write_ctx, case_id, c_result)

        tier = _case_tier(q_result, c_result)
        remaining = _remaining_audit_errors(errors, q_result, c_result)
        if remaining:
            new_case_audit[case_id] = remaining

        side_results = [r for r in (q_result, c_result) if r]
        has_human = any(r["tier"] == "需人工介入" or r["tier"].endswith("V1待处理") for r in side_results)
        has_v2_code = any(r["tier"].endswith("V2待码表更新") for r in side_results)
        fully_resolved = bool(side_results) and not has_human and not has_v2_code
        handled_without_human = bool(side_results) and not has_human

        if fully_resolved:
            stats["resolved"] += 1
        if has_v2_code:
            stats["v2_code_table_update"] += 1
        if has_human:
            stats["needs_human"] += 1

        pending_cases.append({
            "case_id": case_id,
            "scenario": "promo_v2_unmatched",
            "source": "promo",
            "errors": errors,
            "resolved": handled_without_human,
            "tier": tier,
        })

    wb_promo.save(promo_file)
    wb_promo.close()

    now = datetime.now().isoformat()
    state.setdefault("agents", {}).setdefault("fallback", {})
    state["agents"]["fallback"].update({
        "status": "done",
        "completed_at": now,
        "pending_cases": pending_cases,
        "stats": stats,
    })
    state.setdefault("agents", {}).setdefault("promo", {})["fallback_summary"] = {
        "status": "done",
        "completed_at": now,
        **stats,
    }
    state["agents"]["promo"]["case_audit"] = new_case_audit
    _save_state(state_file, state)
    return stats


# ─── 核心入口 ─────────────────────────────────────────────────────────────────

def run_fallback(case_audit: dict, promo_file: str, input_file: str,
                 state: dict, state_file: str | None = None) -> dict:
    """
    促销校验兜底主逻辑。

    执行顺序（每侧独立）：
      竞品侧：先兜底2（截图） → 失败则兜底1（文本）
      Q侧：  兜底1（文本）

    V1（金额平衡）和 V2（码表匹配）独立判断。

    Returns:
        {"fallback1_resolved": int, "fallback2_resolved": int, "needs_human": int}
    """
    if not case_audit:
        _log("兜底：case_audit 为空，跳过")
        return {"fallback1_resolved": 0, "fallback2_resolved": 0, "needs_human": 0}

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        _log("兜底：未检测到 ANTHROPIC_API_KEY，跳过")
        return {"fallback1_resolved": 0, "fallback2_resolved": 0, "needs_human": 0}

    if not promo_file or not os.path.exists(promo_file):
        _log(f"兜底：促销数据文件不存在 {promo_file}，跳过")
        return {"fallback1_resolved": 0, "fallback2_resolved": 0, "needs_human": 0}

    promo_code_set = state.get("input", {}).get("promo_code_set", "xiecheng")
    _log(f"兜底：加载码表（promo_code_set={promo_code_set}）")
    code_q, code_c = _load_code_tables(promo_code_set)

    client = anthropic.Anthropic(api_key=api_key)

    # 读取原始 Excel
    wb_input = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    ws_main = wb_input.worksheets[0]
    raw_headers = [c.value for c in next(ws_main.iter_rows(min_row=1, max_row=1))]
    col_map = {str(h).strip(): i + 1 for i, h in enumerate(raw_headers) if h}

    def get_cell(row_cells, field):
        col = col_map.get(field)
        if not col:
            return ""
        v = row_cells[col - 1].value
        return str(v).strip() if v is not None else ""

    case_rows = {}
    for row in ws_main.iter_rows(min_row=2):
        row_num = row[0].row
        case_no = get_cell(row, CASE_NO_FIELD)
        case_id = case_no if case_no else f"row_{row_num}"
        if case_id in case_audit:
            case_rows[case_id] = {h: get_cell(row, str(h)) for h in raw_headers if h}
            case_rows[case_id]["_row_num"] = row_num

    _log(f"兜底：需处理 {len(case_rows)} 个 case")
    stats = {"fallback1_resolved": 0, "fallback2_resolved": 0, "needs_human": 0}

    for case_id, row_dict in case_rows.items():
        row_num = row_dict["_row_num"]
        _log(f"  case {case_id}（行{row_num}）")

        errors    = case_audit.get(case_id, [])
        has_q_err = any("Q促销" in e or "Q金额" in e for e in errors)
        has_c_err = any("竞品促销" in e or "竞品金额" in e for e in errors)

        q_text    = row_dict.get(Q_PROMO_FIELD, "")
        c_text    = row_dict.get(C_PROMO_FIELD, "")
        image_url = row_dict.get(IMAGE_DETAIL_FIELD, "")

        # 默认：需人工介入
        q_result = {"tier": "需人工介入", "v1_resolved": False, "v2_resolved": False,
                    "enriched": [], "amounts_resolved": None}
        c_result = {"tier": "需人工介入", "v1_resolved": False, "v2_resolved": False,
                    "enriched": [], "amounts_resolved": None}
        q_opt, c_opt = "", ""

        # ── 竞品侧：先兜底2，失败再兜底1 ────────────────────────────────────
        if has_c_err:
            if image_url and image_url.startswith("http"):
                _log(f"    [竞品] 兜底2：Vision 识别截图...")
                fb2 = fallback2_vision(client, image_url)
                if fb2.get("promos"):
                    c_result = _resolve_side(fb2["promos"], row_dict, "c",
                                             code_q, code_c, promo_code_set, "兜底2")
                    c_opt = fb2.get("optimization", "")

            # 兜底2未解决（需人工介入）时再跑兜底1
            if c_result["tier"] == "需人工介入":
                _log(f"    [竞品] 兜底1：LLM 解析文本...")
                fb1_c = fallback1_llm_single(client, c_text)
                if fb1_c.get("promos"):
                    c_result = _resolve_side(fb1_c["promos"], row_dict, "c",
                                             code_q, code_c, promo_code_set, "兜底1")
                    c_opt = fb1_c.get("optimization", "")

        # ── Q侧：兜底1 ──────────────────────────────────────────────────────
        if has_q_err:
            _log(f"    [Q侧] 兜底1：LLM 解析文本...")
            fb1_q = fallback1_llm_single(client, q_text)
            if fb1_q.get("promos"):
                q_result = _resolve_side(fb1_q["promos"], row_dict, "q",
                                         code_q, code_c, promo_code_set, "兜底1")
                q_opt = fb1_q.get("optimization", "")

        # ── 统计 ─────────────────────────────────────────────────────────────
        q_human = has_q_err and q_result["tier"] == "需人工介入"
        c_human = has_c_err and c_result["tier"] == "需人工介入"

        if q_human or c_human:
            stats["needs_human"] += 1

        if has_q_err and q_result["tier"].startswith("兜底1"):
            stats["fallback1_resolved"] += 1
        if has_c_err:
            if c_result["tier"].startswith("兜底2"):
                stats["fallback2_resolved"] += 1
            elif c_result["tier"].startswith("兜底1"):
                stats["fallback1_resolved"] += 1

        _log(f"    → Q侧: {q_result['tier']} (V1={'✅' if q_result['v1_resolved'] else '❌'} V2={'✅' if q_result['v2_resolved'] else '❌'})")
        _log(f"    → C侧: {c_result['tier']} (V1={'✅' if c_result['v1_resolved'] else '❌'} V2={'✅' if c_result['v2_resolved'] else '❌'})")

        # ── 写回促销分行表 ────────────────────────────────────────────────────
        if has_q_err:
            update_promo_file(promo_file, case_id, "q",
                              q_result["tier"], q_opt,
                              q_result["v1_resolved"], q_result["v2_resolved"],
                              q_result["amounts_resolved"])

        if has_c_err:
            update_promo_file(promo_file, case_id, "c",
                              c_result["tier"], c_opt,
                              c_result["v1_resolved"], c_result["v2_resolved"],
                              c_result["amounts_resolved"])

    _log(f"兜底完成：兜底1={stats['fallback1_resolved']} 兜底2={stats['fallback2_resolved']} 人工={stats['needs_human']}")

    if state_file:
        state.setdefault("agents", {}).setdefault("promo", {})["fallback_summary"] = {
            "status": "done",
            "completed_at": datetime.now().isoformat(),
            **stats,
        }
        _save_state(state_file, state)

    return stats


# ─── 独立调试入口 ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--state-file", required=True)
    parser.add_argument("--input-file", required=True)
    parser.add_argument("--llm", action="store_true", help="使用旧版 Claude API/Vision 兜底；默认使用批量脚本兜底")
    args = parser.parse_args()

    state = _load_state(args.state_file)
    case_audit = state.get("agents", {}).get("promo", {}).get("case_audit", {})
    promo_file = state.get("promo_file") or state.get("agents", {}).get("promo", {}).get("promo_file")

    if args.llm:
        stats = run_fallback(case_audit, promo_file, args.input_file, state, args.state_file)
    else:
        stats = run_batch_fallback(state, args.state_file, args.input_file, promo_file)
    print(json.dumps(stats, ensure_ascii=False))


if __name__ == "__main__":
    main()
